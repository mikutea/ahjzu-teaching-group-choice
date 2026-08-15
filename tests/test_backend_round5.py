from __future__ import annotations

import asyncio
import base64
import hashlib
import io
import json
import sqlite3
from dataclasses import replace

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from server.database import connect, initialize_database, utc_now
from server.main import ADMIN_COOKIE, ImportBodyLimitMiddleware
from server.maintenance import check_database, release_check
from server.export_workbook import build_export_workbook
from server.security import (
    activation_code_hash,
    encrypt_activation_code,
    hash_password,
    validate_password_hash,
    verify_password,
)
from server.student_identity import (
    StudentIdentityError,
    normalize_activation_code,
    normalize_student_name,
    normalize_student_number,
)

from .conftest import fictional_document_number


_STRUCTURALLY_VALID_PASSWORD_HASH = "pbkdf2_sha256$600000${}${}".format(
    base64.urlsafe_b64encode(bytes(16)).decode("ascii"),
    base64.urlsafe_b64encode(bytes(32)).decode("ascii"),
)


def roster_csv(rows: list[tuple[str, str, str, str]]) -> bytes:
    lines = ["学号,姓名,专业,证件号"]
    lines.extend(",".join(row) for row in rows)
    return ("\n".join(lines) + "\n").encode("utf-8")


def test_student_entry_page_is_short_lived_edge_cacheable(client: TestClient) -> None:
    student_page = client.get("/")
    assert student_page.status_code == 200
    assert student_page.headers["cache-control"] == "public, max-age=60, must-revalidate"

    assert client.get("/admin").headers["cache-control"] == "no-store"
    assert client.get("/api/public/status").headers["cache-control"] == "no-store"


def import_roster(
    client: TestClient,
    admin_headers: dict[str, str],
    rows: list[tuple[str, str, str, str]],
):
    return client.post(
        "/api/admin/students/import",
        headers=admin_headers,
        files={"files": ("round5.csv", roster_csv(rows), "text/csv")},
    )


def test_roster_import_rejects_unknown_or_duplicate_query_parameters(
    client: TestClient,
    admin_headers: dict[str, str],
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    content = roster_csv(
        [
            (
                "20261234999",
                "严格导入",
                major_name,
                fictional_document_number("strict-query-contract"),
            )
        ]
    )
    for query in (
        "regenerate_existing=true",
        "mode=merge&mode=sync",
    ):
        rejected = client.post(
            f"/api/admin/students/import?{query}",
            headers=admin_headers,
            files={"files": ("round5.csv", content, "text/csv")},
        )
        assert rejected.status_code == 400, rejected.text
        assert "只接受唯一的 mode 参数" in rejected.text
    assert client.get("/api/admin/dashboard").json()["students"] == []


def test_api_surface_contains_only_current_routes_and_models(client: TestClient):
    paths = {
        str(route.path)
        for route in client.app.routes
        if hasattr(route, "path")
    }
    assert "/api/admin/countdown" in paths
    assert not any("start-countdown" in path for path in paths)
    assert not any(
        path.startswith("/api/admin/export/") and path.endswith(".csv")
        for path in paths
    )
    assert not any(path.endswith("/activation-code") for path in paths)
    openapi_text = json.dumps(client.app.openapi(), sort_keys=True)
    assert "source_student_no" not in openapi_text
    assert "regenerate_existing" not in openapi_text


def test_current_identity_contract_is_exact_and_shared_by_roster_import(
    client: TestClient,
    admin_headers: dict[str, str],
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    document_number = fictional_document_number("round5-valid")
    valid_rows = [
        (
            "２０２６１２３４５６７",
            "张 三·Alice",
            major_name,
            document_number,
        )
    ]
    imported = import_roster(client, admin_headers, valid_rows)
    assert imported.status_code == 200, imported.text
    stored = client.get("/api/admin/dashboard").json()["students"]
    assert [(row["student_no"], row["name"]) for row in stored] == [
        ("20261234567", "张 三·Alice")
    ]
    logged_in = client.post(
        "/api/student/login",
        json={
            "student_no": "２０２６１２３４５６７",
            "name": " 张  三·Alice ",
            "activation_code": document_number[-6:].lower(),
        },
    )
    assert logged_in.status_code == 200, logged_in.text
    assert logged_in.json()["student"]["student_no"] == "20261234567"

    assert normalize_student_number("２０２６１２３４５６７") == "20261234567"
    assert normalize_student_name("  Alice   Smith  ") == "Alice Smith"
    assert normalize_activation_code("ａ1ｂ2ｃ3") == "A1B2C3"

    invalid_cases = [
        ("2026123456", "正常姓名", "学号必须是 11 位数字"),
        ("2026123456A", "正常姓名", "学号必须是 11 位数字"),
        ("٢٠٢٦١٢٣٤٥٦٧", "正常姓名", "学号必须是 11 位数字"),
        ("20261234568", "姓名1", "姓名只能包含中文或英文字母"),
        ("20261234569", "张-三", "姓名只能包含中文或英文字母"),
    ]
    for student_no, name, expected in invalid_cases:
        rejected = import_roster(
            client,
            admin_headers,
            [
                (
                    student_no,
                    name,
                    major_name,
                    fictional_document_number(student_no + name),
                )
            ],
        )
        assert rejected.status_code == 400, rejected.text
        assert expected in rejected.json()["detail"]

    with pytest.raises(StudentIdentityError, match="6 位英文字母或数字"):
        normalize_activation_code("A1 B23")


def test_password_hash_validator_accepts_only_current_canonical_structure():
    current = hash_password("Current-Password-Only!")
    assert validate_password_hash(current) is True
    assert verify_password("Current-Password-Only!", current) is True
    assert verify_password("x" * 257, current) is False

    algorithm, iterations, salt, digest = _STRUCTURALLY_VALID_PASSWORD_HASH.split("$")
    invalid_hashes = [
        "",
        f"pbkdf2_sha1${iterations}${salt}${digest}",
        f"{algorithm}$599999${salt}${digest}",
        f"{algorithm}$0600000${salt}${digest}",
        f"{algorithm}${iterations}${salt.rstrip('=')}${digest}",
        f"{algorithm}${iterations}${salt}${digest.rstrip('=')}",
        f"{algorithm}${iterations}$***${digest}",
        f"{algorithm}${iterations}${salt}${digest}$extra",
    ]
    for encoded in invalid_hashes:
        assert validate_password_hash(encoded) is False
        assert verify_password("Current-Password-Only!", encoded) is False

    assert validate_password_hash(hash_password("x" * 12)) is True
    assert validate_password_hash(hash_password("x" * 256)) is True
    for legacy_length in (10, 11):
        legacy_password = "x" * legacy_length
        legacy_salt = bytes(range(16))
        legacy_digest = hashlib.pbkdf2_hmac(
            "sha256", legacy_password.encode("utf-8"), legacy_salt, 600_000
        )
        canonical_legacy_hash = "pbkdf2_sha256$600000${}${}".format(
            base64.urlsafe_b64encode(legacy_salt).decode("ascii"),
            base64.urlsafe_b64encode(legacy_digest).decode("ascii"),
        )
        assert validate_password_hash(canonical_legacy_hash) is True
        assert verify_password(legacy_password, canonical_legacy_hash) is True

    for invalid_password in ("x" * 10, "x" * 11, "x" * 257):
        with pytest.raises(ValueError, match="12 至 256"):
            hash_password(invalid_password)


def test_initialization_rejects_overlong_admin_password(app_config, tmp_path):
    invalid_config = replace(
        app_config,
        database_path=tmp_path / "overlong-admin-password.db",
        admin_initial_password="x" * 257,
    )
    with pytest.raises(ValueError, match="12 至 256"):
        initialize_database(invalid_config)


def _assert_print_ready_table(
    sheet,
    *,
    expected_rows: int,
    expected_columns: int,
) -> None:
    last_column = sheet.cell(row=1, column=expected_columns).column_letter
    assert sheet.max_row == expected_rows
    assert sheet.max_column == expected_columns
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == f"A1:{last_column}{expected_rows}"
    assert sheet.sheet_view.showGridLines is False
    assert sheet.page_setup.orientation == "landscape"
    assert str(sheet.page_setup.paperSize) == str(sheet.PAPERSIZE_A4)
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.page_setup.fitToHeight == 0
    assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert sheet.print_title_rows == "$1:$1"
    assert sheet.print_area == (
        f"'{sheet.title}'!$A$1:${last_column}${expected_rows}"
    )
    assert sheet.print_options.horizontalCentered is True
    assert "安徽建筑大学" in (sheet.oddHeader.center.text or "")
    assert "&P" in (sheet.oddFooter.center.text or "")
    assert "Mikutea" in (sheet.oddFooter.right.text or "")
    assert sheet.page_margins.left <= 0.35
    assert sheet.page_margins.right <= 0.35
    for row in sheet.iter_rows():
        for cell in row:
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"


def test_all_xlsx_exports_are_consistent_wps_print_ready_and_formula_safe(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
):
    dashboard = client.get("/api/admin/dashboard").json()
    major = dashboard["majors"][0]
    group = dashboard["groups"][0]
    student_rows = [
        (
            "20261234571",
            "结果甲",
            major["name"],
            fictional_document_number("round5-export-1"),
        ),
        (
            "20261234572",
            "Result B",
            major["name"],
            fictional_document_number("round5-export-2"),
        ),
    ]
    imported = import_roster(client, admin_headers, student_rows)
    assert imported.status_code == 200, imported.text
    students = {
        row["student_no"]: row
        for row in client.get("/api/admin/dashboard").json()["students"]
    }
    selected = client.post(
        "/api/admin/selections",
        headers=admin_headers,
        json={
            "student_id": students["20261234571"]["id"],
            "group_id": group["id"],
        },
    )
    assert selected.status_code == 200, selected.text

    # Workbook text must remain inert even if an administrator-defined label
    # begins with a spreadsheet formula marker.
    connection = connect(app_config.database_path)
    try:
        now = utc_now()
        connection.execute("UPDATE majors SET name = '=1+1' WHERE id = ?", (major["id"],))
        connection.commit()
    finally:
        connection.close()

    activity_id = admin_headers["X-Activity-ID"]
    complete_response = client.get(
        "/api/admin/export/results.xlsx", params={"activity_id": activity_id}
    )
    selections_response = client.get(
        "/api/admin/export/selections.xlsx", params={"activity_id": activity_id}
    )
    unselected_response = client.get(
        "/api/admin/export/unselected.xlsx", params={"activity_id": activity_id}
    )
    for response in (complete_response, selections_response, unselected_response):
        assert response.status_code == 200, response.text
        assert response.content.startswith(b"PK")
        assert response.headers["cache-control"] == "no-store"
        assert response.headers["x-content-type-options"] == "nosniff"

    complete = load_workbook(io.BytesIO(complete_response.content), data_only=False)
    selections = load_workbook(
        io.BytesIO(selections_response.content), data_only=False
    )
    unselected = load_workbook(
        io.BytesIO(unselected_response.content), data_only=False
    )
    assert complete.sheetnames == ["完整结果", "汇总"]
    assert selections.sheetnames == ["选择记录", "汇总"]
    assert unselected.sheetnames == ["未选名单", "汇总"]
    complete_sheet = complete["完整结果"]
    selection_sheet = selections["选择记录"]
    unselected_sheet = unselected["未选名单"]
    _assert_print_ready_table(
        complete_sheet, expected_rows=3, expected_columns=6
    )
    _assert_print_ready_table(
        selection_sheet, expected_rows=2, expected_columns=6
    )
    _assert_print_ready_table(
        unselected_sheet, expected_rows=2, expected_columns=3
    )

    assert [complete_sheet.cell(row, 1).value for row in (2, 3)] == [
        "20261234571",
        "20261234572",
    ]
    assert selection_sheet["A2"].value == "20261234571"
    assert unselected_sheet["A2"].value == "20261234572"
    assert complete_sheet["C2"].value == "'=1+1"
    assert selection_sheet["C2"].value == "'=1+1"
    assert unselected_sheet["C2"].value == "'=1+1"
    for sheet in (complete_sheet, selection_sheet, unselected_sheet):
        assert sheet["A2"].data_type == "s"
        assert sheet["A2"].number_format == "@"
        assert 14 <= sheet.column_dimensions["A"].width <= 16
    for sheet in (complete_sheet, selection_sheet):
        assert sheet.column_dimensions["F"].width <= 21

    for workbook in (complete, selections, unselected):
        summary = workbook["汇总"]
        assert summary["D8"].value == pytest.approx(0.5)
        assert summary["D8"].number_format == "0.0%"
        assert summary.page_setup.fitToWidth == 1
        assert summary.print_area == f"'汇总'!$A$1:$D${summary.max_row}"
        assert summary.print_options.horizontalCentered is True

    complete.close()
    selections.close()
    unselected.close()


def test_student_credentials_fail_closed_without_activation_ciphertext(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    document_number = fictional_document_number("round5-cipher-required")
    imported = import_roster(
        client,
        admin_headers,
        [("20261234901", "凭据校验", major_name, document_number)],
    )
    assert imported.status_code == 200, imported.text

    connection = connect(app_config.database_path)
    try:
        with pytest.raises(sqlite3.IntegrityError, match="ciphertext required"):
            connection.execute(
                "UPDATE students SET activation_ciphertext = '' WHERE student_no = ?",
                ("20261234901",),
            )
        connection.rollback()

        # Simulate a damaged current-version database without relying on any
        # historical schema fixture. Runtime login and both release validators
        # must reject it rather than accepting a hash-only identity.
        connection.execute("DROP TRIGGER student_activation_ciphertext_guard_insert")
        connection.execute("DROP TRIGGER student_activation_ciphertext_guard_update")
        connection.execute("PRAGMA ignore_check_constraints = ON")
        connection.execute(
            "UPDATE students SET activation_ciphertext = '' WHERE student_no = ?",
            ("20261234901",),
        )
        connection.commit()
    finally:
        connection.close()

    login = client.post(
        "/api/student/login",
        json={
            "student_no": "20261234901",
            "name": "凭据校验",
            "activation_code": document_number[-6:],
        },
    )
    assert login.status_code == 401
    with pytest.raises(RuntimeError, match="缺少激活码密文"):
        check_database(app_config.database_path, app_config.app_secret)
    with pytest.raises(RuntimeError, match="缺少激活码密文"):
        initialize_database(app_config)


def test_dashboard_exposes_strict_activation_code_revealability_boolean(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    document_number = fictional_document_number("round5-revealable")
    imported = import_roster(
        client,
        admin_headers,
        [
            (
                "20261234902",
                "明文查询",
                major_name,
                document_number,
            )
        ],
    )
    assert imported.status_code == 200, imported.text
    student = client.get("/api/admin/dashboard").json()["students"][0]
    assert student["activation_code_revealable"] is True
    assert "activation_ciphertext" not in student
    assert "activation_hash" not in student

    connection = connect(app_config.database_path)
    try:
        connection.execute(
            "UPDATE students SET activation_ciphertext = 'v1.invalid' WHERE student_no = ?",
            ("20261234902",),
        )
        connection.commit()
    finally:
        connection.close()
    corrupted = client.get("/api/admin/dashboard").json()["students"][0]
    assert corrupted["activation_code_revealable"] is False
    login = client.post(
        "/api/student/login",
        json={
            "student_no": "20261234902",
            "name": "明文查询",
            "activation_code": document_number[-6:],
        },
    )
    assert login.status_code == 401
    with pytest.raises(RuntimeError, match="密文无效或与摘要不一致"):
        check_database(app_config.database_path, app_config.app_secret)
    with pytest.raises(RuntimeError, match="密文无效或与摘要不一致"):
        initialize_database(app_config)
    with pytest.raises(RuntimeError, match="密文无效或与摘要不一致"):
        release_check(app_config)


@pytest.mark.parametrize("damage", ["hash", "student_no"])
def test_current_database_rejects_ciphertext_bound_to_wrong_identity(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
    damage: str,
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    document_number = fictional_document_number(f"round5-cipher-{damage}")
    imported = import_roster(
        client,
        admin_headers,
        [("20261234903", "密文绑定校验", major_name, document_number)],
    )
    assert imported.status_code == 200, imported.text

    connection = connect(app_config.database_path)
    try:
        if damage == "hash":
            connection.execute(
                "UPDATE students SET activation_hash = ? WHERE student_no = ?",
                ("0" * 64, "20261234903"),
            )
        else:
            connection.execute(
                "UPDATE students SET student_no = ? WHERE student_no = ?",
                ("20261234904", "20261234903"),
            )
        connection.commit()
    finally:
        connection.close()

    with pytest.raises(RuntimeError, match="密文无效或与摘要不一致"):
        initialize_database(app_config)


def test_untrusted_slow_imports_do_not_consume_authenticated_import_slots(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
):
    session_token = client.cookies.get(ADMIN_COOKIE)
    assert session_token

    async def exercise() -> None:
        downstream_hits: list[list[tuple[bytes, bytes]]] = []

        async def downstream(scope, receive, send) -> None:
            downstream_hits.append(list(scope.get("headers", [])))
            await send(
                {
                    "type": "http.response.start",
                    "status": 200,
                    "headers": [],
                }
            )
            await send({"type": "http.response.body", "body": b""})

        middleware = ImportBodyLimitMiddleware(
            downstream, database_path=app_config.database_path
        )
        release_untrusted = asyncio.Event()

        async def slow_receive():
            await release_untrusted.wait()
            return {"type": "http.request", "body": b"", "more_body": False}

        async def fast_receive():
            return {"type": "http.request", "body": b"", "more_body": False}

        async def ignore_send(message) -> None:
            return None

        base_scope = {
            "type": "http",
            "http_version": "1.1",
            "method": "POST",
            "scheme": "http",
            "path": "/api/admin/students/import",
            "raw_path": b"/api/admin/students/import",
            "query_string": b"",
            "server": ("testserver", 80),
            "client": ("198.51.100.50", 40000),
            "root_path": "",
        }
        untrusted_scope = {**base_scope, "headers": []}
        untrusted_tasks = [
            asyncio.create_task(
                middleware(untrusted_scope, slow_receive, ignore_send)
            )
            for _ in range(4)
        ]
        for _ in range(100):
            if middleware._active_untrusted_imports == 4:
                break
            await asyncio.sleep(0)
        assert middleware._active_untrusted_imports == 4

        trusted_scope = {
            **base_scope,
            "headers": [
                (
                    b"cookie",
                    f"{ADMIN_COOKIE}={session_token}".encode("ascii"),
                )
            ],
        }
        await asyncio.wait_for(
            middleware(trusted_scope, fast_receive, ignore_send), timeout=1
        )
        assert len(downstream_hits) == 1
        assert middleware._active_imports == 0

        release_untrusted.set()
        await asyncio.gather(*untrusted_tasks)
        assert middleware._active_untrusted_imports == 0

    asyncio.run(exercise())


def test_disabled_major_does_not_receive_capacity_when_group_is_resized(
    client: TestClient,
    admin_headers: dict[str, str],
):
    dashboard = client.get("/api/admin/dashboard").json()
    disabled_major = dashboard["majors"][0]
    group = dashboard["groups"][0]
    disabled = client.patch(
        f"/api/admin/majors/{disabled_major['id']}",
        headers=admin_headers,
        json={"active": False},
    )
    assert disabled.status_code == 200, disabled.text

    resized = client.patch(
        f"/api/admin/groups/{group['id']}",
        headers=admin_headers,
        json={"total_capacity": 12},
    )
    assert resized.status_code == 200, resized.text
    refreshed = client.get("/api/admin/dashboard").json()
    group_quotas = [
        quota for quota in refreshed["quotas"] if quota["group_id"] == group["id"]
    ]
    disabled_quota = next(
        quota
        for quota in group_quotas
        if quota["major_id"] == disabled_major["id"]
    )
    active_major_ids = {
        major["id"] for major in refreshed["majors"] if major["active"]
    }
    assert disabled_quota["capacity"] == 0
    assert sum(
        quota["capacity"]
        for quota in group_quotas
        if quota["major_id"] in active_major_ids
    ) == 12


def test_unicode_control_characters_are_rejected_and_xlsx_remains_usable(
    client: TestClient,
    admin_headers: dict[str, str],
):
    invalid_requests = [
        client.patch(
            "/api/admin/settings",
            headers=admin_headers,
            json={"activity_title": "2026级\u200b抢选"},
        ),
        client.post(
            "/api/admin/majors",
            headers=admin_headers,
            json={"name": "建筑\u202e学"},
        ),
        client.post(
            "/api/admin/groups",
            headers=admin_headers,
            json={"name": "教学\u2060组", "total_capacity": 10},
        ),
        client.patch(
            "/api/admin/settings",
            headers=admin_headers,
            json={"public_base_url": "https://class.example/\u200b"},
        ),
    ]
    for response in invalid_requests:
        assert response.status_code == 422, response.text
        assert "控制字符" in response.text

    updated = client.patch(
        "/api/admin/settings",
        headers=admin_headers,
        json={"activity_title": "2026级建筑学教学组抢选"},
    )
    assert updated.status_code == 200, updated.text
    exported = client.get(
        "/api/admin/export/results.xlsx",
        params={"activity_id": admin_headers["X-Activity-ID"]},
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content), data_only=False)
    assert workbook.sheetnames == ["完整结果", "汇总"]
    workbook.close()


@pytest.mark.parametrize(
    "invalid_url",
    [
        "https://",
        "https://choice.example.com bad",
        "https://user:password@choice.example.com",
        "https://choice.example.com/admin",
        "https://choice.example.com/?activity=1",
        "https://choice.example.com/#board",
        "ftp://choice.example.com",
    ],
)
def test_public_base_url_accepts_only_a_canonical_site_origin(
    client: TestClient,
    admin_headers: dict[str, str],
    invalid_url: str,
):
    rejected = client.patch(
        "/api/admin/settings",
        headers=admin_headers,
        json={"public_base_url": invalid_url},
    )
    assert rejected.status_code == 422, rejected.text

    accepted = client.patch(
        "/api/admin/settings",
        headers=admin_headers,
        json={"public_base_url": "HTTPS://CLASS.MIYUO.NET/"},
    )
    assert accepted.status_code == 200, accepted.text
    dashboard = client.get("/api/admin/dashboard").json()
    assert dashboard["settings"]["public_base_url"] == "https://choice.example.com"


def test_selection_can_only_open_through_countdown_endpoint(
    client: TestClient,
    admin_headers: dict[str, str],
):
    rejected = client.post(
        "/api/admin/status",
        headers=admin_headers,
        json={"status": "open"},
    )
    assert rejected.status_code == 409
    assert "10 秒倒计时" in rejected.json()["detail"]
    assert client.get("/api/public/status").json()["status"] == "closed"


def test_release_check_uses_current_database_contract(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
):
    initialize_database(app_config)
    assert release_check(app_config) == "RELEASE_CHECK_OK"


@pytest.mark.parametrize(
    "invalid_hash",
    [
        "",
        _STRUCTURALLY_VALID_PASSWORD_HASH.replace(
            "pbkdf2_sha256", "pbkdf2_sha1", 1
        ),
        _STRUCTURALLY_VALID_PASSWORD_HASH.replace("$600000$", "$599999$", 1),
        "pbkdf2_sha256$600000$not-base64$also-invalid",
    ],
    ids=["blank", "old-algorithm", "old-iterations", "malformed"],
)
def test_current_admin_password_hash_structure_is_fail_closed(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
    invalid_hash: str,
):
    connection = connect(app_config.database_path)
    try:
        connection.execute(
            "UPDATE admin_users SET password_hash = ?",
            (invalid_hash,),
        )
        connection.commit()
    finally:
        connection.close()

    expected = "管理员密码摘要不符合当前 PBKDF2 格式"
    with pytest.raises(RuntimeError, match=expected):
        initialize_database(app_config)
    with pytest.raises(RuntimeError, match=expected):
        check_database(app_config.database_path, app_config.app_secret)
    with pytest.raises(RuntimeError, match=expected):
        release_check(app_config)


@pytest.mark.parametrize("damage", ["student_no", "name"])
def test_current_stored_student_identity_must_be_canonical(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
    damage: str,
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    document_number = fictional_document_number(f"stored-identity-{damage}")
    activation_code = document_number[-6:]
    imported = import_roster(
        client,
        admin_headers,
        [("20261234921", "身份规范", major_name, document_number)],
    )
    assert imported.status_code == 200, imported.text

    connection = connect(app_config.database_path)
    try:
        if damage == "student_no":
            invalid_student_no = "2026123492"
            connection.execute(
                """
                UPDATE students
                SET student_no = ?, activation_hash = ?, activation_ciphertext = ?
                WHERE student_no = ?
                """,
                (
                    invalid_student_no,
                    activation_code_hash(app_config.app_secret, activation_code),
                    encrypt_activation_code(
                        app_config.app_secret, invalid_student_no, activation_code
                    ),
                    "20261234921",
                ),
            )
        else:
            connection.execute(
                "UPDATE students SET name = 'Invalid!Name' WHERE student_no = ?",
                ("20261234921",),
            )
        connection.commit()
    finally:
        connection.close()

    expected = "学生学号或姓名不符合当前规范"
    with pytest.raises(RuntimeError, match=expected):
        initialize_database(app_config)
    with pytest.raises(RuntimeError, match=expected):
        check_database(app_config.database_path, app_config.app_secret)
    with pytest.raises(RuntimeError, match=expected):
        release_check(app_config)


@pytest.mark.parametrize(
    ("mutation_sql", "expected_error"),
    [
        (
            "UPDATE settings SET activity_title = '  非规范标题  ' WHERE id = 1",
            "活动标题不符合当前规范",
        ),
        (
            "UPDATE activities SET code = 'bad code' WHERE status <> 'archived'",
            "活动编码不符合当前规范",
        ),
        (
            "UPDATE majors SET name = '建筑  学' WHERE id = (SELECT MIN(id) FROM majors)",
            "专业名称不符合当前规范",
        ),
        (
            "UPDATE teaching_groups SET name = '教学\u200b组' WHERE id = (SELECT MIN(id) FROM teaching_groups)",
            "教学组名称不符合当前规范",
        ),
        (
            "UPDATE settings SET public_base_url = 'https://class.example/path' WHERE id = 1",
            "访问地址不符合当前规范",
        ),
        (
            "UPDATE admin_users SET username = ' admin '",
            "管理员账号不符合当前规范",
        ),
    ],
)
def test_current_stored_business_text_values_must_be_canonical(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
    mutation_sql: str,
    expected_error: str,
):
    connection = connect(app_config.database_path)
    try:
        connection.execute(mutation_sql)
        connection.commit()
    finally:
        connection.close()

    with pytest.raises(RuntimeError, match=expected_error):
        initialize_database(app_config)
    with pytest.raises(RuntimeError, match=expected_error):
        check_database(app_config.database_path, app_config.app_secret)


@pytest.mark.parametrize(
    ("opened_at", "selection_opens_at", "expected_error"),
    [
        ("garbage", "garbage", "开抢时间格式不正确"),
        (
            "2026-08-14T12:00:00",
            "2026-08-14T12:00:00",
            "开抢时间缺少时区",
        ),
        (
            None,
            "2000-01-01T00:00:00+00:00",
            "开放时间与开抢时间不一致",
        ),
        (
            "2000-01-01T00:00:01+00:00",
            "2000-01-01T00:00:00+00:00",
            "开放时间与开抢时间不一致",
        ),
    ],
)
def test_current_open_activity_timestamps_are_parseable_aware_and_consistent(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
    opened_at: str | None,
    selection_opens_at: str,
    expected_error: str,
):
    connection = connect(app_config.database_path)
    try:
        connection.execute(
            """
            UPDATE activities
            SET status = 'open', opened_at = ?, selection_opens_at = ?
            WHERE status <> 'archived'
            """,
            (opened_at, selection_opens_at),
        )
        connection.commit()
    finally:
        connection.close()

    with pytest.raises(RuntimeError, match=expected_error):
        initialize_database(app_config)
    with pytest.raises(RuntimeError, match=expected_error):
        check_database(app_config.database_path, app_config.app_secret)


@pytest.mark.parametrize(
    ("damage", "expected_error"),
    [
        ("missing_current_activity", "未指向当前活动"),
        ("unassigned_audit", "审计日志未归属活动"),
        ("open_without_countdown", "开放活动缺少开抢时间"),
        ("invalid_countdown_time", "开抢时间格式不正确"),
        ("wrong_copyright", "版权信息与固定发布信息不一致"),
    ],
)
def test_damaged_current_v3_database_is_rejected_without_repair(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
    damage: str,
    expected_error: str,
):
    connection = connect(app_config.database_path)
    try:
        if damage == "missing_current_activity":
            connection.execute(
                "UPDATE settings SET current_activity_id = NULL WHERE id = 1"
            )
        elif damage == "unassigned_audit":
            audit_row = connection.execute(
                """
                INSERT INTO audit_logs
                    (occurred_at, actor_type, actor_id, action, entity_type,
                     entity_id, details_json, activity_id)
                VALUES (?, 'admin', 'test', 'test.damage', 'database',
                        'current', '{}',
                        (SELECT current_activity_id FROM settings WHERE id = 1))
                """,
                (utc_now(),),
            )
            changed = connection.execute(
                "UPDATE audit_logs SET activity_id = NULL WHERE id = ?",
                (audit_row.lastrowid,),
            )
            assert changed.rowcount == 1
        elif damage == "open_without_countdown":
            connection.execute(
                "UPDATE settings SET status = 'open', updated_at = ? WHERE id = 1",
                (utc_now(),),
            )
        elif damage == "invalid_countdown_time":
            connection.execute(
                """
                UPDATE activities
                SET status = 'open', selection_opens_at = 'garbage'
                WHERE status <> 'archived'
                """
            )
        elif damage == "wrong_copyright":
            connection.execute("DROP TRIGGER copyright_settings_guard_update")
            connection.execute("DROP TRIGGER copyright_settings_guard_insert")
            connection.execute(
                "UPDATE settings SET organization_name = '错误版权' WHERE id = 1"
            )
        else:  # pragma: no cover - parameter list is exhaustive
            raise AssertionError(damage)
        connection.commit()
    finally:
        connection.close()

    with pytest.raises(RuntimeError, match=expected_error):
        initialize_database(app_config)

    connection = connect(app_config.database_path)
    try:
        if damage == "missing_current_activity":
            assert connection.execute(
                "SELECT current_activity_id FROM settings WHERE id = 1"
            ).fetchone()[0] is None
        elif damage == "unassigned_audit":
            assert connection.execute(
                "SELECT COUNT(*) FROM audit_logs WHERE activity_id IS NULL"
            ).fetchone()[0] == 1
        elif damage == "open_without_countdown":
            row = connection.execute(
                "SELECT status, selection_opens_at FROM activities WHERE status <> 'archived'"
            ).fetchone()
            assert tuple(row) == ("open", None)
        elif damage == "invalid_countdown_time":
            row = connection.execute(
                "SELECT status, selection_opens_at FROM activities WHERE status <> 'archived'"
            ).fetchone()
            assert tuple(row) == ("open", "garbage")
        elif damage == "wrong_copyright":
            assert connection.execute(
                "SELECT organization_name FROM settings WHERE id = 1"
            ).fetchone()[0] == "错误版权"
    finally:
        connection.close()


@pytest.mark.parametrize(
    ("object_type", "object_name", "expected_error"),
    [
        (
            "TRIGGER",
            "student_activation_ciphertext_guard_insert",
            "缺少关键约束触发器",
        ),
        ("INDEX", "audit_logs_activity_time", "缺少关键索引"),
    ],
)
def test_current_v3_missing_schema_objects_are_rejected_not_recreated(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
    object_type: str,
    object_name: str,
    expected_error: str,
):
    connection = connect(app_config.database_path)
    try:
        connection.execute(f"DROP {object_type} {object_name}")
        connection.commit()
    finally:
        connection.close()

    with pytest.raises(RuntimeError, match=expected_error):
        initialize_database(app_config)
    with pytest.raises(RuntimeError, match=expected_error):
        check_database(app_config.database_path, app_config.app_secret)

    connection = connect(app_config.database_path)
    try:
        assert connection.execute(
            "SELECT 1 FROM sqlite_master WHERE type = ? AND name = ?",
            (object_type.lower(), object_name),
        ).fetchone() is None
    finally:
        connection.close()


@pytest.mark.parametrize(
    ("mutation_sql", "expected_error", "object_type", "object_name"),
    [
        (
            "CREATE TABLE legacy_roster_cache (id INTEGER PRIMARY KEY)",
            "SCHEMA 未定义的结构对象",
            "table",
            "legacy_roster_cache",
        ),
        (
            """
            CREATE TRIGGER legacy_student_update
            AFTER UPDATE ON students
            BEGIN
                SELECT 1;
            END
            """,
            "SCHEMA 未定义的结构对象",
            "trigger",
            "legacy_student_update",
        ),
        (
            """
            DROP TRIGGER selection_capacity_guard;
            CREATE TRIGGER selection_capacity_guard
            BEFORE INSERT ON selections
            BEGIN
                SELECT 1;
            END;
            """,
            "定义与 SCHEMA 不一致",
            "trigger",
            "selection_capacity_guard",
        ),
        (
            """
            DROP INDEX students_major;
            CREATE INDEX students_major ON students(name);
            """,
            "定义与 SCHEMA 不一致",
            "index",
            "students_major",
        ),
        (
            "ALTER TABLE sessions ADD COLUMN legacy_cookie TEXT",
            "定义与 SCHEMA 不一致",
            "table",
            "sessions",
        ),
    ],
)
def test_current_v3_schema_fingerprint_rejects_extra_and_redefined_objects(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
    mutation_sql: str,
    expected_error: str,
    object_type: str,
    object_name: str,
):
    connection = connect(app_config.database_path)
    try:
        connection.executescript(mutation_sql)
    finally:
        connection.close()

    with pytest.raises(RuntimeError, match=expected_error):
        initialize_database(app_config)

    connection = connect(app_config.database_path)
    try:
        assert connection.execute(
            "SELECT 1 FROM sqlite_master WHERE type = ? AND name = ?",
            (object_type, object_name),
        ).fetchone() is not None
    finally:
        connection.close()


def test_current_student_ciphertext_schema_is_nonnullable_and_nonblank(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
):
    connection = connect(app_config.database_path)
    try:
        columns = {
            str(row["name"]): row
            for row in connection.execute("PRAGMA table_info(students)").fetchall()
        }
        assert int(columns["activation_ciphertext"]["notnull"]) == 1
        sql = connection.execute(
            "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'students'"
        ).fetchone()["sql"]
        assert "check(length(trim(activation_ciphertext))>0)" in "".join(
            str(sql).lower().split()
        )
    finally:
        connection.close()


def test_selection_workbook_uses_teacher_facing_source_labels():
    results = [
        {
            "student_no": "20261234573",
            "name": "学生来源",
            "major_name": "建筑学",
            "selection_status": "已选",
            "group_name": "第一教学组",
            "selected_at": "2026-08-14T01:02:03+00:00",
        },
        {
            "student_no": "20261234574",
            "name": "补位来源",
            "major_name": "建筑学",
            "selection_status": "已选",
            "group_name": "第二教学组",
            "selected_at": "2026-08-14T01:03:04+00:00",
        },
    ]
    data = [
        {**results[0], "source": "student"},
        {**results[1], "source": "admin"},
    ]
    payload = build_export_workbook(
        activity={"title": "来源标签测试", "code": "source-labels"},
        exported_at="2026-08-14T01:04:05+00:00",
        kind="selections",
        data_rows=data,
        result_rows=results,
        group_rows=[],
        major_rows=[],
    )
    workbook = load_workbook(io.BytesIO(payload), data_only=False)
    try:
        sheet = workbook["选择记录"]
        assert sheet["F2"].value == "学生提交"
        assert sheet["F3"].value == "管理员补位"
    finally:
        workbook.close()
