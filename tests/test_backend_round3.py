from __future__ import annotations

import io
import json

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pydantic import ValidationError

from server.database import connect
from server.main import StudentLogin
from server.maintenance import check_database
from server.security import activation_code_hash, encrypt_activation_code

from .conftest import TEST_ADMIN_PASSWORD, admin_login, fictional_document_number


def login_headers(client: TestClient) -> dict[str, str]:
    csrf = admin_login(client)
    activity_id = client.get("/api/admin/dashboard").json()["settings"]["activity_id"]
    return {"X-CSRF-Token": csrf, "X-Activity-ID": str(activity_id)}


def import_students(
    client: TestClient,
    headers: dict[str, str],
    rows: list[tuple[str, str, str]],
) -> list[dict]:
    dashboard = client.get("/api/admin/dashboard").json()
    major_by_name = {major["name"]: major for major in dashboard["majors"]}
    csv_rows = ["学号,姓名,专业,证件号"]
    for student_no, name, major_name in rows:
        assert major_name in major_by_name
        csv_rows.append(
            f"{student_no},{name},{major_name},{fictional_document_number(student_no)}"
        )
    response = client.post(
        "/api/admin/students/import",
        headers=headers,
        files={
            "files": (
                "round3.csv",
                ("\n".join(csv_rows) + "\n").encode("utf-8"),
                "text/csv",
            )
        },
    )
    assert response.status_code == 200, response.text
    return client.get("/api/admin/dashboard").json()["students"]


def test_admin_sessions_support_two_devices_logout_and_password_rotation(app):
    computer = TestClient(app)
    phone = TestClient(app)
    forgotten_phone = TestClient(app)
    try:
        computer_csrf = admin_login(computer)
        phone_csrf = admin_login(phone)
        assert computer.get("/api/admin/me").status_code == 200
        assert phone.get("/api/admin/me").status_code == 200

        logged_out = computer.post(
            "/api/admin/logout", headers={"X-CSRF-Token": computer_csrf}
        )
        assert logged_out.status_code == 200, logged_out.text
        assert computer.get("/api/admin/me").status_code == 401
        assert phone.get("/api/admin/me").status_code == 200

        admin_login(forgotten_phone)
        changed = phone.post(
            "/api/admin/password",
            headers={"X-CSRF-Token": phone_csrf},
            json={
                "current_password": TEST_ADMIN_PASSWORD,
                "new_password": "Round3-New-Local-Password!",
            },
        )
        assert changed.status_code == 200, changed.text
        assert phone.get("/api/admin/me").status_code == 200
        assert forgotten_phone.get("/api/admin/me").status_code == 401
    finally:
        computer.close()
        phone.close()
        forgotten_phone.close()


def test_admin_session_count_is_bounded(app, app_config):
    clients = [TestClient(app) for _ in range(9)]
    try:
        for client in clients:
            admin_login(client)
        connection = connect(app_config.database_path)
        try:
            count = connection.execute(
                "SELECT COUNT(*) FROM sessions WHERE role = 'admin'"
            ).fetchone()[0]
        finally:
            connection.close()
        assert count == 8
        assert clients[-1].get("/api/admin/me").status_code == 200
    finally:
        for client in clients:
            client.close()


def test_archive_delete_requires_cas_confirmation_and_leaves_tombstone(
    client: TestClient, admin_headers: dict[str, str], app_config
):
    archived_id = int(admin_headers["X-Activity-ID"])
    created = client.post(
        "/api/admin/activities",
        headers=admin_headers,
        json={
            "title": "第二场后端审计测试",
            "code": "round3-second",
            "copy_structure": True,
            "previous_activity_id": archived_id,
        },
    )
    assert created.status_code == 200, created.text
    current_id = int(created.json()["activity_id"])
    current_headers = {
        "X-CSRF-Token": admin_headers["X-CSRF-Token"],
        "X-Activity-ID": str(current_id),
    }

    wrong_confirmation = client.request(
        "DELETE",
        f"/api/admin/activities/{archived_id}",
        headers=current_headers,
        json={"confirmation": "NO"},
    )
    assert wrong_confirmation.status_code == 422
    stale = client.request(
        "DELETE",
        f"/api/admin/activities/{archived_id}",
        headers={**current_headers, "X-Activity-ID": str(archived_id)},
        json={"confirmation": "DELETE"},
    )
    assert stale.status_code == 409
    current_delete = client.request(
        "DELETE",
        f"/api/admin/activities/{current_id}",
        headers=current_headers,
        json={"confirmation": "DELETE"},
    )
    assert current_delete.status_code == 409

    deleted = client.request(
        "DELETE",
        f"/api/admin/activities/{archived_id}",
        headers=current_headers,
        json={"confirmation": "DELETE"},
    )
    assert deleted.status_code == 200, deleted.text
    assert deleted.json() == {"ok": True, "deleted_activity_id": archived_id}
    activities = client.get("/api/admin/activities").json()
    assert all(int(activity["id"]) != archived_id for activity in activities)
    assert check_database(app_config.database_path, app_config.app_secret) == "ok"

    connection = connect(app_config.database_path)
    try:
        row = connection.execute(
            """
            SELECT activity_id, details_json FROM audit_logs
            WHERE action = 'activity.archive.delete'
            ORDER BY id DESC LIMIT 1
            """
        ).fetchone()
    finally:
        connection.close()
    assert int(row["activity_id"]) == current_id
    details = json.loads(row["details_json"])
    assert details["summary"] == {"students": 0, "selected": 0, "unselected": 0}
    assert "snapshot" not in details


def test_group_capacity_rebalances_quotas_and_honors_selection_floors(
    client: TestClient, admin_headers: dict[str, str]
):
    dashboard = client.get("/api/admin/dashboard").json()
    group = dashboard["groups"][0]
    resized = client.patch(
        f"/api/admin/groups/{group['id']}",
        headers=admin_headers,
        json={"total_capacity": 12},
    )
    assert resized.status_code == 200, resized.text
    assert resized.json()["quotas_adjusted"] is True
    after = client.get("/api/admin/dashboard").json()
    group_quotas = [
        quota for quota in after["quotas"] if quota["group_id"] == group["id"]
    ]
    assert sum(int(quota["capacity"]) for quota in group_quotas) == 12
    assert [int(quota["capacity"]) for quota in group_quotas] == [5, 4, 3]

    majors = after["majors"]
    roster_rows = [
        (f"2026300{index:04d}", f"容量学生{'甲乙丙'[index - 1]}", major["name"])
        for index, major in enumerate(majors, start=1)
    ]
    students = import_students(client, admin_headers, roster_rows)
    student_by_no = {student["student_no"]: student for student in students}
    for student_no, _name, _major in roster_rows:
        assigned = client.post(
            "/api/admin/selections",
            headers=admin_headers,
            json={
                "student_id": student_by_no[student_no]["id"],
                "group_id": group["id"],
            },
        )
        assert assigned.status_code == 200, assigned.text

    too_small = client.patch(
        f"/api/admin/groups/{group['id']}",
        headers=admin_headers,
        json={"total_capacity": 2},
    )
    assert too_small.status_code == 409
    exact = client.patch(
        f"/api/admin/groups/{group['id']}",
        headers=admin_headers,
        json={"total_capacity": 3},
    )
    assert exact.status_code == 200, exact.text
    final = client.get("/api/admin/dashboard").json()
    final_quotas = [
        quota for quota in final["quotas"] if quota["group_id"] == group["id"]
    ]
    assert [int(quota["capacity"]) for quota in final_quotas] == [1, 1, 1]
    assert all(int(quota["capacity"]) >= int(quota["selected_count"]) for quota in final_quotas)


def test_group_rename_or_same_capacity_does_not_rebalance_quotas(
    client: TestClient, admin_headers: dict[str, str]
):
    dashboard = client.get("/api/admin/dashboard").json()
    group = dashboard["groups"][0]
    resized = client.patch(
        f"/api/admin/groups/{group['id']}",
        headers=admin_headers,
        json={"total_capacity": 12},
    )
    assert resized.status_code == 200, resized.text

    first_major = dashboard["majors"][0]
    rows = [
        (f"2026301{index:04d}", f"同容量学生{'甲乙丙'[index - 1]}", first_major["name"])
        for index in range(1, 4)
    ]
    students = import_students(client, admin_headers, rows)
    for student in students:
        assigned = client.post(
            "/api/admin/selections",
            headers=admin_headers,
            json={"student_id": student["id"], "group_id": group["id"]},
        )
        assert assigned.status_code == 200, assigned.text

    before = client.get("/api/admin/dashboard").json()
    before_quotas = {
        int(quota["major_id"]): int(quota["capacity"])
        for quota in before["quotas"]
        if quota["group_id"] == group["id"]
    }
    saved = client.patch(
        f"/api/admin/groups/{group['id']}",
        headers=admin_headers,
        json={"name": "教学组同容量改名", "total_capacity": 12},
    )
    assert saved.status_code == 200, saved.text
    assert saved.json()["quotas_adjusted"] is False
    after = client.get("/api/admin/dashboard").json()
    after_quotas = {
        int(quota["major_id"]): int(quota["capacity"])
        for quota in after["quotas"]
        if quota["group_id"] == group["id"]
    }
    assert after_quotas == before_quotas


def test_xlsx_export_contains_complete_roster_and_wps_safe_student_numbers(
    client: TestClient, admin_headers: dict[str, str]
):
    dashboard = client.get("/api/admin/dashboard").json()
    major_name = dashboard["majors"][0]["name"]
    rows = [
        ("20263020001", "结果甲", major_name),
        ("20263020002", "结果乙", major_name),
    ]
    students = import_students(client, admin_headers, rows)
    selected = client.post(
        "/api/admin/selections",
        headers=admin_headers,
        json={"student_id": students[0]["id"], "group_id": dashboard["groups"][0]["id"]},
    )
    assert selected.status_code == 200, selected.text

    exported = client.get(
        "/api/admin/export/results.xlsx",
        params={"activity_id": admin_headers["X-Activity-ID"]},
    )
    assert exported.status_code == 200, exported.text
    assert exported.content.startswith(b"PK")
    workbook = load_workbook(io.BytesIO(exported.content), data_only=False)
    assert workbook.sheetnames == ["完整结果", "汇总"]
    results = workbook["完整结果"]
    assert [cell.value for cell in results[1]] == [
        "学号",
        "姓名",
        "专业",
        "状态",
        "教学组",
        "选择时间",
    ]
    assert results["A2"].value == rows[0][0]
    assert results["A2"].data_type == "s"
    assert results["A2"].number_format == "@"
    assert results.column_dimensions["A"].width >= 14
    assert results.freeze_panes == "A2"
    assert results.auto_filter.ref == "A1:F3"
    assert {results["D2"].value, results["D3"].value} == {"已选", "未选"}
    summary = workbook["汇总"]
    assert summary["A1"].value == "教学组抢选结果汇总"
    assert summary["D8"].value == pytest.approx(0.5)
    assert summary["D8"].number_format == "0.0%"


def test_xlsx_export_neutralizes_formula_like_roster_text(
    client: TestClient, admin_headers: dict[str, str], app_config
):
    dashboard = client.get("/api/admin/dashboard").json()
    major = dashboard["majors"][0]
    connection = connect(app_config.database_path)
    try:
        connection.execute("BEGIN IMMEDIATE")
        connection.execute("UPDATE majors SET name = '=1+1' WHERE id = ?", (major["id"],))
        now = "2026-08-13T00:00:00+00:00"
        activation_code = fictional_document_number("formula-roster")[-6:]
        connection.execute(
            """
            INSERT INTO students
                (student_no, name, major_id, activation_hash, activation_ciphertext,
                 active, created_at, updated_at)
            VALUES ('20263029999', '@SUM(A1:A2)', ?, ?, ?, 1, ?, ?)
            """,
            (
                major["id"],
                activation_code_hash(app_config.app_secret, activation_code),
                encrypt_activation_code(
                    app_config.app_secret, "20263029999", activation_code
                ),
                now,
                now,
            ),
        )
        connection.commit()
    finally:
        connection.close()
    exported = client.get(
        "/api/admin/export/results.xlsx",
        params={"activity_id": admin_headers["X-Activity-ID"]},
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content), data_only=False)
    results = workbook["完整结果"]
    assert results["A2"].value == "20263029999"
    assert results["A2"].data_type == "s"
    assert results["B2"].value == "'@SUM(A1:A2)"
    assert results["B2"].data_type == "s"
    assert results["C2"].value == "'=1+1"
    assert results["C2"].data_type == "s"


@pytest.mark.parametrize(
    "payload",
    [
        {"student_no": "12", "name": "合法姓名", "activation_code": "123456"},
        {"student_no": "2026<script>", "name": "合法姓名", "activation_code": "123456"},
        {"student_no": "20260000001", "name": "<script>", "activation_code": "123456"},
        {"student_no": "20260000001", "name": "姓名🙂", "activation_code": "123456"},
        {"student_no": "20260000001", "name": "姓名\n注入", "activation_code": "123456"},
        {"student_no": "20260000001", "name": "合法姓名", "activation_code": "12-456"},
        {"student_no": "20260000001", "name": "合法姓名", "activation_code": "1234567"},
    ],
)
def test_student_login_rejects_illegal_fields(payload: dict[str, str]):
    with pytest.raises(ValidationError):
        StudentLogin.model_validate(payload)


@pytest.mark.parametrize("name", ["欧阳·子涵", "Chen Wei Lun", "陳・美玲", "王小明"])
def test_student_login_accepts_supported_mainland_hk_macao_taiwan_names(name: str):
    model = StudentLogin.model_validate(
        {"student_no": "20261234567", "name": name, "activation_code": "１２３ＡＢＣ"}
    )
    assert model.activation_code == "123ABC"
