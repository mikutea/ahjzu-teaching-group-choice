from __future__ import annotations

import io

from openpyxl import load_workbook

from server.export_workbook import build_export_workbook


def test_long_current_names_wrap_without_truncation_in_print_workbook():
    activity_title = ("建筑与空间规划学院教学组抢选活动" * 6)[:80]
    student_name = ("张李王赵周吴郑冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜" * 2)[:40]
    major_name = ("城乡规划与区域发展综合实验专业方向" * 6)[:80]
    group_name = ("城乡空间设计与数字技术融合实践教学组" * 6)[:80]
    selected_at = "2026-08-14T09:30:00+08:00"
    result_rows = [
        {
            "student_no": "20261234567",
            "name": student_name,
            "major_name": major_name,
            "selection_status": "已选",
            "group_name": group_name,
            "selected_at": selected_at,
        }
    ]

    payload = build_export_workbook(
        activity={"title": activity_title, "code": "round5-print"},
        exported_at="2026-08-14T10:00:00+08:00",
        kind="complete",
        data_rows=result_rows,
        result_rows=result_rows,
        group_rows=[
            {
                "name": group_name,
                "total_capacity": 30,
                "selected_count": 1,
            }
        ],
        major_rows=[
            {
                "name": major_name,
                "student_count": 1,
                "selected_count": 1,
            }
        ],
    )

    workbook = load_workbook(io.BytesIO(payload), data_only=False)
    try:
        sheet = workbook["完整结果"]
        assert sheet["B2"].value == student_name
        assert sheet["C2"].value == major_name
        assert sheet["E2"].value == group_name
        assert all(sheet.cell(2, column).alignment.horizontal == "center" for column in range(1, 7))
        assert all(sheet.cell(2, column).alignment.vertical == "center" for column in range(1, 7))
        assert all(sheet.cell(2, column).alignment.wrap_text is True for column in (2, 3, 5))
        assert sheet.row_dimensions[2].height > 22
        assert 14 <= sheet.column_dimensions["A"].width <= 16
        assert sheet.column_dimensions["B"].width <= 20
        assert sheet.column_dimensions["C"].width <= 26
        assert sheet.column_dimensions["E"].width <= 26
        assert sheet.page_setup.fitToWidth == 1
        assert sheet.print_options.horizontalCentered is True

        summary = workbook["汇总"]
        assert summary["B3"].value == activity_title
        assert summary["B3"].alignment.wrap_text is True
        assert summary.row_dimensions[3].height > 22
        assert summary["A12"].value == group_name
        assert summary["A12"].alignment.wrap_text is True
        assert summary.row_dimensions[12].height > 22
        assert summary["A16"].value == major_name
        assert summary["A16"].alignment.wrap_text is True
        assert summary.row_dimensions[16].height > 22
    finally:
        workbook.close()
