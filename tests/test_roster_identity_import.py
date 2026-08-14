from __future__ import annotations

import csv
import io
import subprocess
from html.parser import HTMLParser
from pathlib import Path

import pytest
import xlwt
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from server.database import connect
from server.roster import RosterParseError, activation_code_from_document_number

from .conftest import fictional_document_number, open_selection_now


def roster_csv(rows: list[tuple[str, str, str, str]]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["证件号", "姓名", "专业名称", "学号"])
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def roster_xlsx(rows: list[tuple[object, str, str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "名单"
    sheet.append(["证件号", "姓名", "专业名称", "学号"])
    for row in rows:
        sheet.append(row)
    decoy = workbook.create_sheet("不应读取")
    decoy.append(["错误表头"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def roster_xls(rows: list[tuple[object, str, str, str]]) -> bytes:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("名单")
    for column, heading in enumerate(["证件号", "姓名", "专业名称", "学号"]):
        sheet.write(0, column, heading)
    for row_index, row in enumerate(rows, start=1):
        for column, value in enumerate(row):
            sheet.write(row_index, column, value)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def upload_files(
    client: TestClient,
    headers: dict[str, str],
    uploads: list[tuple[str, bytes, str]],
    *,
    mode: str = "merge",
):
    return client.post(
        "/api/admin/students/import",
        headers=headers,
        params={"mode": mode},
        files=[("files", upload) for upload in uploads],
    )


def assert_identifiers_absent(payload: bytes | str, identifiers: list[str]) -> None:
    text = payload.decode("utf-8", errors="ignore") if isinstance(payload, bytes) else payload
    for identifier in identifiers:
        assert identifier not in text


def test_csv_xls_xlsx_sync_is_one_combined_roster_and_identity_never_persists(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
):
    created_major = client.post(
        "/api/admin/majors",
        headers=admin_headers,
        json={"name": "Test Major"},
    )
    assert created_major.status_code == 200, created_major.text

    documents = {
        "20265000001": "H00000000000123456",
        "20265000002": "H00000000000234567",
        "20265000003": "H00000000000345678",
        "20265000004": "H00000000000456789",
    }
    initial = client.post(
        "/api/admin/students/import",
        headers=admin_headers,
        files={
            "files": (
                "initial.csv",
                roster_csv(
                    [
                        (documents["20265000001"], "Fictional CSV", "Test Major", "20265000001"),
                        (documents["20265000002"], "Fictional XLS", "Test Major", "20265000002"),
                        (documents["20265000003"], "Fictional XLSX", "Test Major", "20265000003"),
                        (documents["20265000004"], "Fictional Omitted", "Test Major", "20265000004"),
                    ]
                ),
                "text/csv",
            )
        },
    )
    assert initial.status_code == 200, initial.text
    assert initial.json() == {
        "created": 4,
        "updated": 0,
        "deactivated": 0,
        "rotated": 0,
        "file_count": 1,
        "row_count": 4,
        "majors_created": [],
        "majors_reactivated": [],
        "activation_code_policy": "normalized_document_number_last_6",
    }

    synchronized = upload_files(
        client,
        admin_headers,
        [
            (
                "one.csv",
                roster_csv(
                    [(documents["20265000001"], "Fictional CSV", "Test Major", "20265000001")]
                ),
                "text/csv",
            ),
            (
                "two.xls",
                roster_xls(
                    [
                        (
                            documents["20265000002"],
                            "Fictional XLS",
                            "Test Major",
                            "20265000002",
                        )
                    ]
                ),
                "application/vnd.ms-excel",
            ),
            (
                "three.xlsx",
                roster_xlsx(
                    [(documents["20265000003"], "Fictional XLSX", "Test Major", "20265000003")]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ],
        mode="sync",
    )
    assert synchronized.status_code == 200, synchronized.text
    assert synchronized.json() == {
        "created": 0,
        "updated": 3,
        "deactivated": 1,
        "rotated": 0,
        "file_count": 3,
        "row_count": 3,
        "majors_created": [],
        "majors_reactivated": [],
        "activation_code_policy": "normalized_document_number_last_6",
    }
    assert "credentials" not in synchronized.json()
    assert '"activation_code":' not in synchronized.text
    assert_identifiers_absent(initial.text + synchronized.text, list(documents.values()))

    dashboard = client.get("/api/admin/dashboard").json()
    active_by_number = {
        row["student_no"]: bool(row["active"])
        for row in dashboard["students"]
    }
    assert active_by_number == {
        "20265000001": True,
        "20265000002": True,
        "20265000003": True,
        "20265000004": False,
    }
    xls_student = next(
        row for row in dashboard["students"] if row["student_no"] == "20265000002"
    )
    revealed = client.post(
        f"/api/admin/students/{xls_student['id']}/activation-code/reveal",
        headers=admin_headers,
    )
    assert revealed.status_code == 200, revealed.text
    assert revealed.json()["credential"]["activation_code"] == "234567"
    assert_identifiers_absent(revealed.text, list(documents.values()))

    audit = client.get("/api/admin/audit")
    assert audit.status_code == 200, audit.text
    assert_identifiers_absent(audit.text, list(documents.values()))
    assert_identifiers_absent(app_config.database_path.read_bytes(), list(documents.values()))

    old_activity = int(dashboard["settings"]["activity_id"])
    archived = client.post(
        "/api/admin/activities",
        headers=admin_headers,
        json={
            "title": "Identity Redaction Next Activity",
            "code": "identity-redaction-next",
            "copy_structure": True,
            "previous_activity_id": old_activity,
        },
    )
    assert archived.status_code == 200, archived.text
    archive = client.get(f"/api/admin/activities/{old_activity}/archive.json")
    assert archive.status_code == 200, archive.text
    assert_identifiers_absent(archive.content, list(documents.values()))


@pytest.mark.parametrize(
    ("document_number", "expected_code"),
    [
        ("H00000000000654321", "654321"),
        ("00000020000101001X", "01001X"),
        ("Z000000(A)", "00000A"),
        ("1000000(1)", "000001"),
        ("A100000009", "000009"),
        ("AB00000008", "000008"),
    ],
)
def test_supported_document_formats_derive_last_six_and_can_login_after_open(
    client: TestClient,
    admin_headers: dict[str, str],
    document_number: str,
    expected_code: str,
):
    dashboard = client.get("/api/admin/dashboard").json()
    major_name = dashboard["majors"][0]["name"]
    suffix = expected_code.replace("X", "9").replace("A", "8")
    student_no = f"20265{suffix}"
    imported = client.post(
        "/api/admin/students/import",
        headers=admin_headers,
        files={
            "files": (
                "format.csv",
                roster_csv([(document_number, "虚构格式学生", major_name, student_no)]),
                "text/csv",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    assert "credentials" not in imported.json()
    assert document_number not in imported.text

    open_selection_now(client, admin_headers)
    public_status = client.get("/api/public/status")
    assert public_status.status_code == 200, public_status.text
    assert public_status.json()["student_login_allowed"] is True
    assert "仍可登录" in public_status.json()["status_message"]

    student = TestClient(client.app)
    try:
        login = student.post(
            "/api/student/login",
            json={
                "student_no": student_no,
                "name": "虚构格式学生",
                "activation_code": expected_code,
            },
        )
        assert login.status_code == 200, login.text
        assert login.json()["phase"] == "open"
        assert login.json()["student_login_allowed"] is True
        assert "仍可登录" in login.json()["status_message"]
    finally:
        student.close()


@pytest.mark.parametrize(
    "document_number",
    [
        "١٢٣٤٥٦٧٨٩٠١٢٣٤٥",
        "A١٢٣٤٥٦(٧)",
        "1١٢٣٤٥٦٧",
        "H١٢٣٤٥٦٧٨",
        "A1١٢٣٤٥٦٧٨",
        "T١٢٣٤٥٦٧٨٩٠١٢٣٤٥٦X",
    ],
)
def test_document_numbers_reject_non_ascii_decimal_digits(document_number: str):
    with pytest.raises(RosterParseError, match="证件号格式无法识别"):
        activation_code_from_document_number(document_number)


@pytest.mark.parametrize(
    ("document_number", "expected_code"),
    [
        ("３４０１０２２００６０１０１１２３４", "011234"),
        ("Ａ１２３４５６（７）", "1234567"[-6:]),
        ("Ｈ１２３４５６７８", "345678"),
    ],
)
def test_document_numbers_nfkc_fullwidth_digits_to_ascii(
    document_number: str, expected_code: str
):
    code = activation_code_from_document_number(document_number)
    assert code == expected_code
    assert code.isascii()


def test_complete_results_export_contains_every_active_student(
    client: TestClient,
    admin_headers: dict[str, str],
):
    dashboard = client.get("/api/admin/dashboard").json()
    activity_id = int(dashboard["settings"]["activity_id"])
    major_name = dashboard["majors"][0]["name"]
    group = dashboard["groups"][0]
    documents = [fictional_document_number("result-selected"), fictional_document_number("result-waiting")]
    imported = client.post(
        "/api/admin/students/import",
        headers=admin_headers,
        files={
            "files": (
                "results.csv",
                roster_csv(
                    [
                        (documents[0], "虚构已选学生", major_name, "20265010001"),
                        (documents[1], "虚构未选学生", major_name, "20265010002"),
                    ]
                ),
                "text/csv",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    students = {
        row["student_no"]: row
        for row in client.get("/api/admin/dashboard").json()["students"]
    }
    assigned = client.post(
        "/api/admin/selections",
        headers=admin_headers,
        json={"student_id": students["20265010001"]["id"], "group_id": group["id"]},
    )
    assert assigned.status_code == 200, assigned.text

    exported = client.get(
        "/api/admin/export/results.xlsx",
        params={"activity_id": activity_id},
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content), read_only=True, data_only=True)
    rows = list(workbook["完整结果"].iter_rows(values_only=True))
    workbook.close()
    assert rows[0] == ("学号", "姓名", "专业", "状态", "教学组", "选择时间")
    by_number = {
        row[0]: dict(zip(rows[0], row, strict=True))
        for row in rows[1:]
    }
    assert set(by_number) == {"20265010001", "20265010002"}
    assert by_number["20265010001"]["状态"] == "已选"
    assert by_number["20265010001"]["教学组"] == group["name"]
    assert by_number["20265010001"]["选择时间"]
    assert by_number["20265010002"]["状态"] == "未选"
    assert by_number["20265010002"]["教学组"] is None
    assert by_number["20265010002"]["选择时间"] is None
    assert_identifiers_absent(exported.content, documents)


@pytest.mark.parametrize("cross_file", [False, True], ids=["same-file", "cross-file"])
def test_duplicate_student_number_rolls_back_the_whole_batch(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
    cross_file: bool,
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    documents = [
        fictional_document_number("duplicate-one"),
        fictional_document_number("duplicate-two"),
        fictional_document_number("unique-one"),
        fictional_document_number("unique-two"),
    ]
    first_rows = [
        (documents[2], "虚构唯一甲", major_name, "20265020001"),
        (documents[0], "虚构重复甲", major_name, "20265029999"),
    ]
    second_rows = [
        (documents[3], "虚构唯一乙", major_name, "20265020002"),
        (documents[1], "虚构重复乙", major_name, "20265029999"),
    ]
    uploads = (
        [
            ("first.csv", roster_csv(first_rows), "text/csv"),
            ("second.csv", roster_csv(second_rows), "text/csv"),
        ]
        if cross_file
        else [("same.csv", roster_csv(first_rows + second_rows), "text/csv")]
    )
    rejected = upload_files(client, admin_headers, uploads)
    assert rejected.status_code == 400, rejected.text
    assert "重复" in rejected.json()["detail"]
    assert client.get("/api/admin/dashboard").json()["students"] == []
    connection = connect(app_config.database_path)
    try:
        assert connection.execute("SELECT COUNT(*) FROM students").fetchone()[0] == 0
    finally:
        connection.close()
    assert_identifiers_absent(app_config.database_path.read_bytes(), documents)


def test_missing_document_cannot_generate_an_activation_code(
    client: TestClient,
    admin_headers: dict[str, str],
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    missing = client.post(
        "/api/admin/students/import",
        headers=admin_headers,
        files={
            "files": (
                "missing-document.csv",
                (
                    "学号,姓名,专业名称\n"
                    f"20265030001,虚构缺失证件,{major_name}\n"
                ).encode("utf-8"),
                "text/csv",
            )
        },
    )
    assert missing.status_code == 400, missing.text
    assert "缺少必要表头" in missing.json()["detail"]
    assert client.get("/api/admin/dashboard").json()["students"] == []


def test_reimported_document_suffix_rotates_derived_credential_and_revokes_session(
    client: TestClient,
    admin_headers: dict[str, str],
    app_config,
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    old_document = "H00000000000111111"
    new_document = "H00000000000222222"

    first = client.post(
        "/api/admin/students/import",
        headers=admin_headers,
        files={
            "files": (
                "first.csv",
                roster_csv([(old_document, "虚构换证学生", major_name, "20265030003")]),
                "text/csv",
            )
        },
    )
    assert first.status_code == 200, first.text

    student = TestClient(client.app)
    try:
        login = student.post(
            "/api/student/login",
            json={
                "student_no": "20265030003",
                "name": "虚构换证学生",
                "activation_code": "111111",
            },
        )
        assert login.status_code == 200, login.text

        changed = client.post(
            "/api/admin/students/import",
            headers=admin_headers,
            files={
                "files": (
                    "changed.csv",
                    roster_csv([(new_document, "虚构换证学生", major_name, "20265030003")]),
                    "text/csv",
                )
            },
        )
        assert changed.status_code == 200, changed.text
        assert changed.json()["updated"] == 1
        assert changed.json()["rotated"] == 1
        assert '"activation_code":' not in changed.text
        assert student.get("/api/student/me").status_code == 401

        old_code = TestClient(client.app)
        new_code = TestClient(client.app)
        try:
            assert old_code.post(
                "/api/student/login",
                json={
                    "student_no": "20265030003",
                    "name": "虚构换证学生",
                    "activation_code": "111111",
                },
            ).status_code == 401
            assert new_code.post(
                "/api/student/login",
                json={
                    "student_no": "20265030003",
                    "name": "虚构换证学生",
                    "activation_code": "222222",
                },
            ).status_code == 200
        finally:
            old_code.close()
            new_code.close()
    finally:
        student.close()

    assert_identifiers_absent(app_config.database_path.read_bytes(), [old_document, new_document])
    audit = client.get("/api/admin/audit")
    assert audit.status_code == 200, audit.text
    assert_identifiers_absent(audit.text, [old_document, new_document])


def test_xlsx_long_numeric_document_cell_is_rejected_as_precision_risk(
    client: TestClient,
    admin_headers: dict[str, str],
):
    major_name = client.get("/api/admin/dashboard").json()["majors"][0]["name"]
    rejected = upload_files(
        client,
        admin_headers,
        [
            (
                "numeric.xlsx",
                roster_xlsx([(123456789012345678, "虚构数值单元格", major_name, "20265040001")]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        ],
    )
    assert rejected.status_code == 400, rejected.text
    assert "必须设置为文本" in rejected.json()["detail"]
    assert client.get("/api/admin/dashboard").json()["students"] == []


class _DomContractParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.by_id: dict[str, dict[str, str | None]] = {}

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        element_id = attributes.get("id")
        if element_id:
            self.by_id[element_id] = {"tag": tag, **attributes}


def test_projection_and_roster_ui_dom_contract_and_guidance_copy():
    root = Path(__file__).resolve().parents[1]
    admin_html = (root / "web" / "admin.html").read_text(encoding="utf-8")
    admin_js = (root / "web" / "admin.js").read_text(encoding="utf-8")
    app_css = (root / "web" / "app.css").read_text(encoding="utf-8")
    parser = _DomContractParser()
    parser.feed(admin_html)

    assert parser.by_id["live-selection-feed"]["tag"] == "div"
    assert parser.by_id["export-complete-results"]["tag"] == "a"
    file_input = parser.by_id["student-csv"]
    assert file_input["tag"] == "input"
    assert "multiple" in file_input
    assert all(extension in str(file_input["accept"]) for extension in (".csv", ".xls", ".xlsx"))
    assert "regenerate-existing" not in parser.by_id
    assert "download-last-credentials" not in parser.by_id
    assert admin_html.count('id="danger-dialog"') == 1
    live_board_block = admin_html.split('<div id="live-board"', 1)[1].split(
        '<section class="content-card recent-card"', 1
    )[0]
    assert '<dialog id="danger-dialog"' not in live_board_block
    assert "const fullscreenHost = document.fullscreenElement" in admin_js
    assert "fullscreenHost.append(adminEls.dangerDialog)" in admin_js
    assert "originalParent.append(adminEls.dangerDialog)" in admin_js
    board_logo_css = app_css.split(".college-wordmark--board {", 1)[1].split(".board-presentation-brand h1", 1)[0]
    shared_logo_css = app_css.split(".college-wordmark img {", 1)[1].split(".college-wordmark--mobile", 1)[0]
    assert "background: transparent" in board_logo_css
    assert "position: absolute" in shared_logo_css
    assert "width: 108.4623%" in shared_logo_css

    assert "实时选择流水" in admin_html
    assert "尚未进入候场" in admin_html
    assert "连续滚动" in admin_js
    assert "支持批量选择 CSV / XLS / XLSX" in admin_html
    assert "个人激活码固定为证件号后 6 位" in admin_html
    assert "系统不会随机生成或批量重置" in admin_html
    assert "导出本场完整结果" in admin_html
    assert "扫码仍可登录 · 名额与名单实时更新" in admin_js
    template_block = admin_js.split('document.querySelector("#download-template")', 1)[1].split("for (const [anchor", 1)[0]
    assert '[["学号", "姓名", "专业名称", "证件号"]]' in template_block
    assert "示例学生" not in template_block


def test_repository_privacy_guards_ignore_all_roster_files():
    root = Path(__file__).resolve().parents[1]
    gitignore = (root / ".gitignore").read_text(encoding="utf-8")
    dockerignore = (root / ".dockerignore").read_text(encoding="utf-8")

    assert "*.csv" in gitignore
    assert "!examples/" not in gitignore
    assert "*.xls" in gitignore and "*.xlsx" in gitignore
    assert "*.csv" in dockerignore
    assert "*.xls" in dockerignore and "*.xlsx" in dockerignore

    ignored = subprocess.run(
        ["git", "check-ignore", "--no-index", "server/fictional-sensitive-roster.csv"],
        cwd=root,
        capture_output=True,
        text=True,
        check=False,
    )
    assert ignored.returncode == 0, ignored.stderr
