from __future__ import annotations

import csv
import io
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from itertools import islice
from pathlib import Path
from typing import Any, Iterable

import xlrd
from openpyxl import load_workbook

from .student_identity import (
    StudentIdentityError,
    normalize_student_name,
    normalize_student_number,
)


MAX_XLSX_MEMBERS = 512
MAX_XLSX_UNCOMPRESSED_BYTES = 16 * 1024 * 1024
MAX_TABLE_ROWS = 2_101
MAX_TABLE_COLUMNS = 64

_DOCUMENT_SEPARATORS = re.compile(r"[\s()（）/\\\-‐‑‒–—―]+")
_DOCUMENT_PATTERNS = (
    # Mainland resident identity cards (current and legacy).
    re.compile(r"^\d{17}[0-9X]$"),
    re.compile(r"^\d{15}$"),
    # Hong Kong identity card, Macao resident identity card, and Taiwan ID /
    # unified certificate formats. Formatting punctuation is removed first.
    re.compile(r"^[A-Z]{1,2}\d{6}[0-9A]$"),
    re.compile(r"^[157]\d{7}$"),
    re.compile(r"^[HM]\d{8}$"),
    re.compile(r"^\d{8}$"),
    re.compile(r"^[A-Z][1289]\d{8}$"),
    re.compile(r"^[A-Z]{2}\d{8}$"),
    # School registry expansion: H, M, or T followed by 17 positions.
    re.compile(r"^[HMT]\d{16}[0-9X]$"),
)

HEADER_ALIASES = {
    "student_no": ("student_no", "学号"),
    "name": ("name", "姓名"),
    "major": ("major", "专业", "专业名称"),
    "document_number": (
        "document_number",
        "document_no",
        "id_number",
        "id_no",
        "证件号",
        "身份证号",
        "证件号码",
        "身份证号码",
    ),
}


class RosterParseError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedRosterRow:
    file_index: int
    line_number: int
    student_no: str
    name: str
    major_name: str
    activation_code: str


def _clean_text(value: str) -> str:
    return " ".join(value.strip().split())


def activation_code_from_document_number(value: str) -> str:
    """Derive a credential without retaining or returning the source identifier."""

    normalized = unicodedata.normalize("NFKC", value).upper()
    normalized = _DOCUMENT_SEPARATORS.sub("", normalized)
    if not any(pattern.fullmatch(normalized) for pattern in _DOCUMENT_PATTERNS):
        raise RosterParseError("证件号格式无法识别")
    if len(normalized) < 6:
        raise RosterParseError("证件号规范化后少于 6 位")
    return normalized[-6:]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean_text(str(value))


def _identity_cell_text(value: Any) -> str:
    """Preserve raw text so the shared identity validator sees control chars."""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _document_cell_text(value: Any) -> str:
    if isinstance(value, bool):
        raise RosterParseError("证件号单元格类型不正确")
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            raise RosterParseError("证件号单元格必须设置为文本")
        digits = str(int(value))
        if len(digits) > 15:
            raise RosterParseError("超过 15 位的证件号单元格必须设置为文本")
        return digits
    return _clean_text("" if value is None else str(value))


def _bounded_rows(rows: Iterable[Iterable[Any]]) -> list[list[Any]]:
    bounded: list[list[Any]] = []
    for row_number, raw_row in enumerate(rows, start=1):
        if row_number > MAX_TABLE_ROWS:
            raise RosterParseError("名单文件行数过多")
        row = list(raw_row)
        if len(row) > MAX_TABLE_COLUMNS and any(
            _cell_text(value) for value in row[MAX_TABLE_COLUMNS:]
        ):
            raise RosterParseError("名单文件列数过多")
        bounded.append(row[:MAX_TABLE_COLUMNS])
    return bounded


def _csv_rows(content: bytes) -> list[list[Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("gb18030")
        except UnicodeDecodeError as exc:
            raise RosterParseError("CSV 编码需为 UTF-8 或 GB18030") from exc
    try:
        return _bounded_rows(csv.reader(io.StringIO(text)))
    except csv.Error as exc:
        raise RosterParseError("CSV 格式无法解析") from exc


def _preflight_xlsx(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise RosterParseError("XLSX 文件包含过多内部成员")
            if any(member.flag_bits & 0x1 for member in members):
                raise RosterParseError("不支持加密的 XLSX 文件")
            if sum(member.file_size for member in members) > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise RosterParseError("XLSX 解压后内容过大")
    except zipfile.BadZipFile as exc:
        raise RosterParseError("XLSX 文件结构无法解析") from exc


def _xlsx_rows(content: bytes) -> list[list[Any]]:
    _preflight_xlsx(content)
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        raise RosterParseError("XLSX 文件无法解析") from exc
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(
            min_row=1, max_col=MAX_TABLE_COLUMNS + 1, values_only=True
        )
        return _bounded_rows(islice(rows, MAX_TABLE_ROWS + 1))
    finally:
        workbook.close()


def _xls_rows(content: bytes) -> list[list[Any]]:
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        if workbook.nsheets < 1:
            raise RosterParseError("XLS 文件不包含工作表")
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows > MAX_TABLE_ROWS:
            raise RosterParseError("名单文件行数过多")
        if sheet.ncols > MAX_TABLE_COLUMNS:
            for row_index in range(sheet.nrows):
                if any(
                    _cell_text(value)
                    for value in sheet.row_values(
                        row_index, start_colx=MAX_TABLE_COLUMNS
                    )
                ):
                    raise RosterParseError("名单文件列数过多")
        return [
            sheet.row_values(row_index, end_colx=MAX_TABLE_COLUMNS)
            for row_index in range(sheet.nrows)
        ]
    except RosterParseError:
        raise
    except Exception as exc:
        raise RosterParseError("XLS 文件无法解析") from exc
    finally:
        if "workbook" in locals():
            workbook.release_resources()


def _table_rows(filename: str, content: bytes) -> list[list[Any]]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        return _csv_rows(content)
    if suffix == ".xlsx":
        if not content.startswith(b"PK"):
            raise RosterParseError("XLSX 文件签名不正确")
        return _xlsx_rows(content)
    if suffix == ".xls":
        if not content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
            raise RosterParseError("XLS 文件签名不正确")
        return _xls_rows(content)
    raise RosterParseError("仅支持 CSV、XLS 和 XLSX 文件")


def _first_nonempty_row(rows: list[list[Any]]) -> tuple[int, list[Any]]:
    for index, row in enumerate(rows):
        if any(_cell_text(value) for value in row):
            return index, row
    raise RosterParseError("名单文件为空")


def parse_roster_file(
    *, filename: str, content: bytes, file_index: int
) -> list[ParsedRosterRow]:
    rows = _table_rows(filename, content)
    header_index, raw_header = _first_nonempty_row(rows)
    alias_lookup = {
        alias.casefold(): key
        for key, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }
    header_indexes: dict[str, int] = {}
    for index, raw_value in enumerate(raw_header):
        header = _clean_text(unicodedata.normalize("NFKC", _cell_text(raw_value))).casefold()
        key = alias_lookup.get(header)
        if not key:
            continue
        if key in header_indexes:
            raise RosterParseError(f"表头“{HEADER_ALIASES[key][-1]}”重复")
        header_indexes[key] = index
    missing = [
        HEADER_ALIASES[key][-1]
        for key in ("student_no", "name", "major", "document_number")
        if key not in header_indexes
    ]
    if missing:
        raise RosterParseError("缺少必要表头：" + "、".join(missing))

    parsed_rows: list[ParsedRosterRow] = []
    for row_index, raw_row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(_cell_text(value) for value in raw_row):
            continue
        values: dict[str, str] = {}
        for key, index in header_indexes.items():
            cell = raw_row[index] if index < len(raw_row) else None
            if key == "document_number":
                values[key] = _document_cell_text(cell)
            elif key in {"student_no", "name"}:
                values[key] = _identity_cell_text(cell)
            else:
                values[key] = _cell_text(cell)
        if not values["student_no"] or not values["name"] or not values["major"]:
            raise RosterParseError(f"第 {row_index} 行必须填写学号、姓名和专业")
        if not values["document_number"]:
            raise RosterParseError(f"第 {row_index} 行必须填写证件号")
        try:
            student_no = normalize_student_number(values["student_no"])
            name = normalize_student_name(values["name"])
            activation_code = activation_code_from_document_number(
                values["document_number"]
            )
        except (RosterParseError, StudentIdentityError) as exc:
            raise RosterParseError(f"第 {row_index} 行{exc}") from exc
        parsed_rows.append(
            ParsedRosterRow(
                file_index=file_index,
                line_number=row_index,
                student_no=student_no,
                name=name,
                major_name=values["major"],
                activation_code=activation_code,
            )
        )
    if not parsed_rows:
        raise RosterParseError("名单文件没有学生记录")
    return parsed_rows


def parse_roster_files(
    uploads: Iterable[tuple[str, bytes]],
) -> list[ParsedRosterRow]:
    parsed: list[ParsedRosterRow] = []
    for file_index, (filename, content) in enumerate(uploads, start=1):
        try:
            parsed.extend(
                parse_roster_file(
                    filename=filename, content=content, file_index=file_index
                )
            )
        except RosterParseError as exc:
            raise RosterParseError(f"第 {file_index} 个文件：{exc}") from exc
    return parsed
