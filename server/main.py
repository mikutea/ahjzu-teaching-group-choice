from __future__ import annotations

import asyncio
import csv
import hashlib
import hmac
import io
import json
import re
import secrets
import sqlite3
import threading
import time
import unicodedata
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta, timezone
from ipaddress import ip_address
from pathlib import Path
from typing import Annotated, Any, Literal

import qrcode
from fastapi import FastAPI, File, HTTPException, Request, Response, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from pydantic import BaseModel, Field, ValidationInfo, field_validator
from starlette.types import ASGIApp, Message, Receive, Scope, Send

from .config import Config, PROJECT_ROOT
from .database import (
    SCHEMA_VERSION,
    activity_snapshot,
    audit,
    connect,
    initialize_database,
    utc_now,
)
from .roster import RosterParseError, parse_roster_files
from .security import (
    activation_code_hash,
    decrypt_activation_code,
    encrypt_activation_code,
    hash_password,
    new_csrf_token,
    new_session_token,
    session_token_hash,
    verify_activation_code,
    verify_password,
)


WEB_ROOT = PROJECT_ROOT / "web"
BRAND_ROOT = PROJECT_ROOT / "assets" / "brand"
ADMIN_COOKIE = "tg_admin_session"
STUDENT_COOKIE = "tg_student_session"
MAX_ROSTER_FILE_BYTES = 1_048_576
MAX_IMPORT_BODY_BYTES = 1_250_000
MAX_IMPORT_FILES = 12
MAX_ROSTER_ROWS = 2_000
MAX_CONCURRENT_IMPORTS = 4
IMPORT_BODY_TIMEOUT_SECONDS = 20
FIXED_ORGANIZATION_NAME = "安徽建筑大学 · 建筑与空间规划学院"
FIXED_OWNER_NAME = "Mikutea"
COUNTDOWN_SECONDS = 10
PRESENCE_FRESH_SECONDS = 35
HEARTBEAT_WRITE_INTERVAL_SECONDS = 15
MAX_ADMIN_SESSIONS_PER_USER = 8

STUDENT_NUMBER_PATTERN = re.compile(r"^[A-Za-z0-9_-]{4,32}$")
STUDENT_NAME_PATTERN = re.compile(
    r"^[A-Za-z0-9\u00C0-\u024F\u1E00-\u1EFF"
    r"\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF"
    r"\U00020000-\U0002FA1F ·•・'’\-‐‑]+$"
)
ACTIVATION_CODE_PATTERN = re.compile(r"^[A-Z0-9]{6}$")


def session_utc_now() -> str:
    """Wall clock for credential expiry, intentionally separate from event test clocks."""

    return datetime.now(UTC).isoformat(timespec="seconds")


class ImportBodyLimitMiddleware:
    """Reject oversized roster uploads before multipart parsing begins."""

    def __init__(self, app: ASGIApp) -> None:
        self.app = app
        self._active_imports = 0
        self._gate_lock = threading.Lock()

    async def __call__(self, scope: Scope, receive: Receive, send: Send) -> None:
        if scope["type"] != "http" or scope.get("path") != "/api/admin/students/import":
            await self.app(scope, receive, send)
            return

        rejected = False
        with self._gate_lock:
            if self._active_imports >= MAX_CONCURRENT_IMPORTS:
                rejected = True
            else:
                self._active_imports += 1
        if rejected:
            await self._send_error(
                send,
                429,
                "当前正在处理其他名单，请稍后重试",
                extra_headers=[(b"retry-after", b"1")],
            )
            return

        try:
            await self._handle_import(scope, receive, send)
        finally:
            with self._gate_lock:
                self._active_imports -= 1

    async def _handle_import(
        self, scope: Scope, receive: Receive, send: Send
    ) -> None:

        headers = {key.lower(): value for key, value in scope.get("headers", [])}
        content_length = headers.get(b"content-length")
        if content_length:
            try:
                declared_length = int(content_length)
            except ValueError:
                await self._send_error(send, 400, "请求长度格式不正确")
                return
            if declared_length > MAX_IMPORT_BODY_BYTES:
                await self._send_error(send, 413, "上传请求不能超过 1.25 MB")
                return

        received = 0
        chunks: list[bytes] = []
        deadline = time.monotonic() + IMPORT_BODY_TIMEOUT_SECONDS
        try:
            while True:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    raise TimeoutError
                message = await asyncio.wait_for(receive(), timeout=remaining)
                if message["type"] == "http.disconnect":
                    return
                body = message.get("body", b"")
                received += len(body)
                if received > MAX_IMPORT_BODY_BYTES:
                    await self._send_error(send, 413, "上传请求不能超过 1.25 MB")
                    return
                chunks.append(body)
                if not message.get("more_body", False):
                    break
        except TimeoutError:
            await self._send_error(send, 408, "名单上传超时，请检查网络后重试")
            return

        replayed = False

        async def replay_receive() -> Message:
            nonlocal replayed
            if not replayed:
                replayed = True
                return {
                    "type": "http.request",
                    "body": b"".join(chunks),
                    "more_body": False,
                }
            return {"type": "http.request", "body": b"", "more_body": False}

        await self.app(scope, replay_receive, send)

    @staticmethod
    async def _send_error(
        send: Send,
        status: int,
        detail: str,
        *,
        extra_headers: list[tuple[bytes, bytes]] | None = None,
    ) -> None:
        body = json.dumps({"detail": detail}, ensure_ascii=False).encode("utf-8")
        headers = [
            (b"content-type", b"application/json; charset=utf-8"),
            (b"content-length", str(len(body)).encode("ascii")),
            (b"cache-control", b"no-store"),
        ]
        headers.extend(extra_headers or [])
        await send(
            {
                "type": "http.response.start",
                "status": status,
                "headers": headers,
            }
        )
        await send({"type": "http.response.body", "body": body})


def clean_text(value: str) -> str:
    return " ".join(value.strip().split())


def resolve_client_host(request: Request, trusted_proxy_ips: tuple[str, ...]) -> str:
    direct_host = request.client.host if request.client else "unknown"
    if direct_host not in trusted_proxy_ips:
        return direct_host

    cloudflare_host = request.headers.get("CF-Connecting-IP", "").strip()
    try:
        return str(ip_address(cloudflare_host))
    except ValueError:
        return direct_host


class StrictModel(BaseModel):
    model_config = {"extra": "forbid"}


class AdminLogin(StrictModel):
    username: str = Field(min_length=1, max_length=80)
    password: str = Field(min_length=1, max_length=256)


class StudentLogin(StrictModel):
    student_no: str = Field(min_length=4, max_length=32)
    name: str = Field(min_length=1, max_length=80)
    activation_code: str = Field(min_length=6, max_length=6)

    @field_validator("student_no", "name", mode="before")
    @classmethod
    def strip_values(cls, value: Any, info: ValidationInfo) -> Any:
        if not isinstance(value, str):
            return value
        if any(unicodedata.category(character).startswith("C") for character in value):
            raise ValueError(f"{info.field_name} 不能包含控制字符")
        return clean_text(value)

    @field_validator("student_no")
    @classmethod
    def validate_student_number(cls, value: str) -> str:
        if not STUDENT_NUMBER_PATTERN.fullmatch(value):
            raise ValueError("学号只能包含英文字母、数字、下划线或连字符")
        return value

    @field_validator("name")
    @classmethod
    def validate_student_name(cls, value: str) -> str:
        if not STUDENT_NAME_PATTERN.fullmatch(value):
            raise ValueError("姓名包含不支持的字符")
        return value

    @field_validator("activation_code", mode="before")
    @classmethod
    def normalize_activation_code(cls, value: Any) -> Any:
        if not isinstance(value, str):
            return value
        normalized = "".join(
            unicodedata.normalize("NFKC", value).strip().upper().split()
        )
        if not ACTIVATION_CODE_PATTERN.fullmatch(normalized):
            raise ValueError("个人激活码必须是 6 位英文字母或数字")
        return normalized


class StatusUpdate(StrictModel):
    status: Literal["closed", "open"]


class SettingsUpdate(StrictModel):
    activity_title: str | None = Field(default=None, min_length=2, max_length=120)
    public_base_url: str | None = Field(default=None, max_length=300)

    @field_validator("activity_title", mode="before")
    @classmethod
    def normalize_names(cls, value: Any) -> Any:
        return clean_text(value) if isinstance(value, str) else value

    @field_validator("public_base_url")
    @classmethod
    def validate_base_url(cls, value: str | None) -> str | None:
        if value is None:
            return None
        value = value.strip().rstrip("/")
        if value and not value.startswith(("http://", "https://")):
            raise ValueError("访问地址必须以 http:// 或 https:// 开头")
        return value


class ActivityCreate(StrictModel):
    title: str = Field(min_length=2, max_length=120)
    code: str | None = Field(default=None, min_length=2, max_length=48, pattern=r"^[A-Za-z0-9_-]+$")
    copy_structure: bool = True
    previous_activity_id: int = Field(gt=0)

    @field_validator("title", mode="before")
    @classmethod
    def normalize_title(cls, value: Any) -> Any:
        return clean_text(value) if isinstance(value, str) else value


class MajorCreate(StrictModel):
    name: str = Field(min_length=1, max_length=80)

    @field_validator("name", mode="before")
    @classmethod
    def normalize_name(cls, value: Any) -> Any:
        return clean_text(value) if isinstance(value, str) else value


class MajorUpdate(StrictModel):
    name: str | None = Field(default=None, min_length=1, max_length=80)
    active: bool | None = None
    sort_order: int | None = Field(default=None, ge=0, le=100_000)

    @field_validator("name", mode="before")
    @classmethod
    def normalize_name(cls, value: Any) -> Any:
        return clean_text(value) if isinstance(value, str) else value


class GroupCreate(StrictModel):
    name: str = Field(min_length=1, max_length=80)
    total_capacity: int = Field(default=30, ge=0, le=1000)

    @field_validator("name", mode="before")
    @classmethod
    def normalize_name(cls, value: Any) -> Any:
        return clean_text(value) if isinstance(value, str) else value


class GroupUpdate(StrictModel):
    name: str | None = Field(default=None, min_length=1, max_length=80)
    total_capacity: int | None = Field(default=None, ge=0, le=1000)
    active: bool | None = None
    sort_order: int | None = Field(default=None, ge=0, le=100_000)

    @field_validator("name", mode="before")
    @classmethod
    def normalize_name(cls, value: Any) -> Any:
        return clean_text(value) if isinstance(value, str) else value


class QuotaUpdate(StrictModel):
    capacity: int = Field(ge=0, le=1000)


class StudentSelect(StrictModel):
    group_id: int = Field(gt=0)


class AdminAssign(StrictModel):
    student_id: int = Field(gt=0)
    group_id: int = Field(gt=0)


class RevokeSelection(StrictModel):
    student_id: int = Field(gt=0)
    reason: str = Field(default="管理员撤销", min_length=2, max_length=200)

    @field_validator("reason")
    @classmethod
    def normalize_reason(cls, value: str) -> str:
        return clean_text(value)


class PasswordChange(StrictModel):
    current_password: str = Field(min_length=1, max_length=256)
    new_password: str = Field(min_length=12, max_length=256)


class ArchiveDelete(StrictModel):
    confirmation: Literal["DELETE"]


@dataclass(frozen=True)
class Identity:
    role: Literal["student", "admin"]
    subject_id: int
    csrf_token: str
    token_hash: str


class RateLimiter:
    def __init__(self, *, max_keys: int = 10_000) -> None:
        self._events: dict[str, list[float]] = {}
        self._lock = threading.Lock()
        self._max_keys = max_keys
        self._last_cleanup = 0.0

    def check(self, key: str, *, limit: int, window_seconds: int) -> None:
        now = time.monotonic()
        cutoff = now - window_seconds
        with self._lock:
            if now - self._last_cleanup >= 60:
                self._events = {
                    existing_key: [event for event in events if event >= cutoff]
                    for existing_key, events in self._events.items()
                    if any(event >= cutoff for event in events)
                }
                self._last_cleanup = now
            if key not in self._events and len(self._events) >= self._max_keys:
                oldest_key = min(
                    self._events,
                    key=lambda existing_key: self._events[existing_key][-1],
                )
                self._events.pop(oldest_key, None)
            events = [event for event in self._events.get(key, []) if event >= cutoff]
            if len(events) >= limit:
                raise HTTPException(status_code=429, detail="尝试次数过多，请稍后再试")
            events.append(now)
            self._events[key] = events

    def clear(self, key: str) -> None:
        with self._lock:
            self._events.pop(key, None)


def create_app(config: Config | None = None) -> FastAPI:
    config = config or Config.from_env()
    initialize_database(config)
    limiter = RateLimiter()
    app = FastAPI(
        title="教学组抢选系统",
        docs_url=None if config.environment == "production" else "/api/docs",
        redoc_url=None,
        openapi_url=None if config.environment == "production" else "/api/openapi.json",
    )
    app.state.config = config
    app.add_middleware(ImportBodyLimitMiddleware)

    @app.middleware("http")
    async def security_headers(request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "same-origin"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; img-src 'self' data:; style-src 'self'; "
            "script-src 'self'; connect-src 'self'; frame-ancestors 'none'; "
            "base-uri 'self'; form-action 'self'"
        )
        if config.cookie_secure or config.public_base_url.startswith("https://"):
            response.headers["Strict-Transport-Security"] = "max-age=31536000"
        if request.url.path.startswith("/api/") or request.url.path in {"/", "/admin"}:
            if not response.headers.get("Cache-Control"):
                response.headers["Cache-Control"] = "no-store"
        return response

    @app.exception_handler(sqlite3.IntegrityError)
    async def sqlite_integrity_error(_: Request, exc: sqlite3.IntegrityError):
        message = str(exc).lower()
        if "unique" in message:
            detail = "名称或编号已存在"
        else:
            detail = "数据约束校验失败"
        return JSONResponse(status_code=409, content={"detail": detail})

    @app.exception_handler(sqlite3.OperationalError)
    async def sqlite_operational_error(_: Request, exc: sqlite3.OperationalError):
        message = str(exc).lower()
        if "locked" in message or "busy" in message:
            return JSONResponse(
                status_code=503,
                content={"detail": "当前提交人数较多，请稍候重试"},
                headers={"Retry-After": "1"},
            )
        raise exc

    def client_key(request: Request, namespace: str) -> str:
        host = resolve_client_host(request, config.trusted_proxy_ips)
        return f"{namespace}:{host}"

    def principal_key(namespace: str, value: str) -> str:
        digest = hmac.new(
            config.app_secret.encode("utf-8"),
            f"{namespace}:{clean_text(value).casefold()}".encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        return f"{namespace}:{digest[:24]}"

    def require_session_from_connection(
        request: Request,
        role: Literal["student", "admin"],
        connection: sqlite3.Connection,
        *,
        csrf: bool = False,
    ) -> Identity:
        cookie_name = ADMIN_COOKIE if role == "admin" else STUDENT_COOKIE
        token = request.cookies.get(cookie_name, "")
        if not token:
            raise HTTPException(status_code=401, detail="请先登录")
        token_hash = session_token_hash(token)
        row = connection.execute(
            """
            SELECT role, subject_id, csrf_token
            FROM sessions
            WHERE token_hash = ? AND expires_at > ?
            """,
            (token_hash, session_utc_now()),
        ).fetchone()
        if not row or row["role"] != role:
            raise HTTPException(status_code=401, detail="登录已失效，请重新登录")
        if csrf:
            provided = request.headers.get("X-CSRF-Token", "")
            if not secrets.compare_digest(provided, row["csrf_token"]):
                raise HTTPException(status_code=403, detail="请求校验失败，请刷新页面后重试")
        return Identity(
            role=role,
            subject_id=int(row["subject_id"]),
            csrf_token=row["csrf_token"],
            token_hash=token_hash,
        )

    def require_session(
        request: Request,
        role: Literal["student", "admin"],
        *,
        csrf: bool = False,
    ) -> Identity:
        connection = connect(config.database_path)
        try:
            return require_session_from_connection(request, role, connection, csrf=csrf)
        finally:
            connection.close()

    def revalidate_identity(
        request: Request,
        connection: sqlite3.Connection,
        identity: Identity,
    ) -> Identity:
        current = require_session_from_connection(
            request, identity.role, connection, csrf=True
        )
        if (
            current.subject_id != identity.subject_id
            or not secrets.compare_digest(current.token_hash, identity.token_hash)
        ):
            raise HTTPException(status_code=401, detail="登录已失效，请重新登录")
        return current

    def persist_session(
        connection: sqlite3.Connection,
        *,
        role: Literal["student", "admin"],
        subject_id: int,
        token: str,
        csrf_token: str,
        now_dt: datetime,
    ) -> None:
        expires = now_dt + timedelta(hours=config.session_hours)
        connection.execute(
            "DELETE FROM sessions WHERE expires_at <= ?", (session_utc_now(),)
        )
        if role == "student":
            connection.execute(
                "DELETE FROM sessions WHERE role = 'student' AND subject_id = ?",
                (subject_id,),
            )
        connection.execute(
            """
            INSERT INTO sessions
                (token_hash, role, subject_id, csrf_token, created_at,
                 last_seen_at, expires_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session_token_hash(token),
                role,
                subject_id,
                csrf_token,
                now_dt.isoformat(timespec="seconds"),
                now_dt.isoformat(timespec="seconds"),
                expires.isoformat(timespec="seconds"),
            ),
        )
        if role == "admin":
            connection.execute(
                """
                DELETE FROM sessions
                WHERE token_hash IN (
                    SELECT token_hash FROM sessions
                    WHERE role = 'admin' AND subject_id = ? AND token_hash <> ?
                    ORDER BY created_at DESC, token_hash DESC
                    LIMIT -1 OFFSET ?
                )
                """,
                (
                    subject_id,
                    session_token_hash(token),
                    MAX_ADMIN_SESSIONS_PER_USER - 1,
                ),
            )

    def set_session_cookie(response: Response, role: Literal["student", "admin"], token: str) -> None:
        cookie_name = ADMIN_COOKIE if role == "admin" else STUDENT_COOKIE
        response.set_cookie(
            cookie_name,
            token,
            max_age=config.session_hours * 3600,
            httponly=True,
            secure=config.cookie_secure,
            samesite="strict",
            path="/",
        )

    def current_activity(connection: sqlite3.Connection) -> sqlite3.Row:
        activity = connection.execute(
            """
            SELECT a.* FROM activities a
            JOIN settings s ON s.current_activity_id = a.id
            WHERE s.id = 1
            """
        ).fetchone()
        if not activity:
            raise RuntimeError("当前活动不存在")
        return activity

    def parse_utc(value: str) -> datetime:
        parsed = datetime.fromisoformat(value)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=UTC)
        return parsed.astimezone(UTC)

    def activity_phase(activity: sqlite3.Row, *, server_now: str | None = None) -> str:
        if activity["status"] != "open":
            return (
                "waiting"
                if activity["status"] == "closed" and activity["opened_at"] is None
                else "closed"
            )
        opens_at = activity["selection_opens_at"]
        if opens_at and parse_utc(server_now or utc_now()) < parse_utc(str(opens_at)):
            return "countdown"
        return "open"

    def student_status_message(phase: str) -> str:
        messages = {
            "waiting": "学生登录已开放，抢选尚未开始。",
            "countdown": "学生登录已开放，统一倒计时进行中，请等待开抢。",
            "open": "抢选进行中，学生仍可登录并提交选择。",
            "closed": "本场抢选已结束，学生仍可登录查看选择结果。",
        }
        return messages.get(phase, "学生仍可登录查看当前活动状态。")

    def ensure_closed(connection: sqlite3.Connection) -> None:
        status = current_activity(connection)["status"]
        if status != "closed":
            raise HTTPException(status_code=409, detail="请先关闭抢选再修改结构或配额")

    def rebalance_group_quotas(
        connection: sqlite3.Connection, group_id: int, total_capacity: int
    ) -> list[dict[str, Any]]:
        """Deterministically fit the major matrix to a group's new capacity.

        Existing quotas are the preferred weights.  When every quota is zero,
        active roster counts become the weights, with an equal active-major
        fallback.  Current selections are hard lower bounds.  Integer leftovers
        use the largest-remainder method with stable major ordering.
        """

        now = utc_now()
        connection.execute(
            """
            INSERT OR IGNORE INTO quotas (major_id, group_id, capacity, updated_at)
            SELECT id, ?, 0, ? FROM majors
            """,
            (group_id, now),
        )
        rows = connection.execute(
            """
            SELECT q.major_id, q.capacity, m.name AS major_name, m.active,
                   m.sort_order,
                   (SELECT COUNT(*) FROM students roster
                    WHERE roster.major_id = q.major_id AND roster.active = 1
                   ) AS student_count,
                   (SELECT COUNT(*) FROM selections se
                    JOIN students selected_student
                      ON selected_student.id = se.student_id
                    WHERE se.group_id = q.group_id
                      AND selected_student.major_id = q.major_id
                      AND se.revoked_at IS NULL
                   ) AS selected_count
            FROM quotas q JOIN majors m ON m.id = q.major_id
            WHERE q.group_id = ?
            ORDER BY m.sort_order, m.id
            """,
            (group_id,),
        ).fetchall()
        selected_total = sum(int(row["selected_count"]) for row in rows)
        if total_capacity < selected_total:
            raise HTTPException(
                status_code=409,
                detail=f"总容量不能小于当前已选人数 {selected_total}",
            )
        if not rows:
            connection.execute(
                "UPDATE teaching_groups SET total_capacity = ?, updated_at = ? WHERE id = ?",
                (total_capacity, now, group_id),
            )
            return []

        weights = [int(row["capacity"]) for row in rows]
        if not any(weights):
            weights = [
                int(row["student_count"]) if bool(row["active"]) else 0
                for row in rows
            ]
        if not any(weights):
            weights = [1 if bool(row["active"]) else 0 for row in rows]
        if not any(weights):
            weights = [1 for _ in rows]

        remaining = total_capacity - selected_total
        weight_total = sum(weights)
        shares = [remaining * weight // weight_total for weight in weights]
        remainders = [remaining * weight % weight_total for weight in weights]
        undistributed = remaining - sum(shares)
        remainder_order = sorted(
            range(len(rows)),
            key=lambda index: (
                -remainders[index],
                int(rows[index]["sort_order"]),
                int(rows[index]["major_id"]),
            ),
        )
        for index in remainder_order[:undistributed]:
            shares[index] += 1
        targets = [
            int(row["selected_count"]) + shares[index]
            for index, row in enumerate(rows)
        ]

        # Lower to selection floors before changing total capacity.  This
        # ordering remains valid for both shrinking and growing under SQLite's
        # quota/group guard triggers; final values are invisible until commit.
        for row in rows:
            connection.execute(
                """
                UPDATE quotas SET capacity = ?, updated_at = ?
                WHERE major_id = ? AND group_id = ?
                """,
                (int(row["selected_count"]), now, int(row["major_id"]), group_id),
            )
        connection.execute(
            "UPDATE teaching_groups SET total_capacity = ?, updated_at = ? WHERE id = ?",
            (total_capacity, now, group_id),
        )
        adjustments: list[dict[str, Any]] = []
        for row, target in zip(rows, targets, strict=True):
            connection.execute(
                """
                UPDATE quotas SET capacity = ?, updated_at = ?
                WHERE major_id = ? AND group_id = ?
                """,
                (target, now, int(row["major_id"]), group_id),
            )
            if int(row["capacity"]) != target:
                adjustments.append(
                    {
                        "major_id": int(row["major_id"]),
                        "major_name": row["major_name"],
                        "from": int(row["capacity"]),
                        "to": target,
                        "selected_count": int(row["selected_count"]),
                    }
                )
        return adjustments

    def require_expected_activity(
        request: Request,
        connection: sqlite3.Connection,
        *,
        identity: Identity,
    ) -> sqlite3.Row:
        revalidate_identity(request, connection, identity)
        supplied = request.headers.get("X-Activity-ID", "").strip()
        if not supplied:
            raise HTTPException(status_code=428, detail="缺少活动版本，请刷新页面后重试")
        try:
            expected_id = int(supplied)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="活动版本格式不正确") from exc
        activity = current_activity(connection)
        if int(activity["id"]) != expected_id:
            raise HTTPException(status_code=409, detail="当前活动已经变化，请刷新页面后重试")
        return activity

    def setting_dict(connection: sqlite3.Connection) -> dict[str, Any]:
        row = connection.execute(
            """
            SELECT s.*, a.id AS activity_id, a.code AS activity_code,
                   a.title AS current_activity_title, a.status AS activity_status,
                   a.created_at AS activity_created_at,
                   a.opened_at AS opened_at,
                   a.selection_opens_at AS selection_opens_at
            FROM settings s JOIN activities a ON a.id = s.current_activity_id
            WHERE s.id = 1
            """
        ).fetchone()
        server_now = utc_now()
        phase = activity_phase(row, server_now=server_now)
        return {
            "activity_id": row["activity_id"],
            "activity_code": row["activity_code"],
            "activity_title": row["current_activity_title"],
            "activity_created_at": row["activity_created_at"],
            "organization_name": FIXED_ORGANIZATION_NAME,
            "owner_name": FIXED_OWNER_NAME,
            "status": row["activity_status"],
            "phase": phase,
            "student_login_allowed": True,
            "status_message": student_status_message(phase),
            "server_now": server_now,
            "selection_opens_at": row["selection_opens_at"],
            "public_base_url": row["public_base_url"],
            "updated_at": row["updated_at"],
        }

    def activity_readiness(connection: sqlite3.Connection) -> dict[str, Any]:
        blockers: list[str] = []
        warnings: list[str] = []
        settings = setting_dict(connection)
        student_total = int(
            connection.execute("SELECT COUNT(*) FROM students WHERE active = 1").fetchone()[0]
        )
        group_total = int(
            connection.execute(
                "SELECT COUNT(*) FROM teaching_groups WHERE active = 1"
            ).fetchone()[0]
        )
        if student_total == 0:
            blockers.append("尚未导入有效学生名单")
        if group_total == 0:
            blockers.append("尚未启用任何教学组")

        inactive_major_students = int(
            connection.execute(
                """
                SELECT COUNT(*) FROM students s
                JOIN majors m ON m.id = s.major_id
                WHERE s.active = 1 AND m.active = 0
                """
            ).fetchone()[0]
        )
        if inactive_major_students:
            blockers.append(f"有 {inactive_major_students} 名学生所属专业已停用")

        active_capacity = int(
            connection.execute(
                "SELECT COALESCE(SUM(total_capacity), 0) FROM teaching_groups WHERE active = 1"
            ).fetchone()[0]
        )
        if student_total and active_capacity < student_total:
            blockers.append(
                f"启用教学组总容量仅 {active_capacity}，少于有效学生 {student_total} 人"
            )

        for row in connection.execute(
            """
            SELECT g.name, g.total_capacity, COALESCE(SUM(q.capacity), 0) AS quota_total
            FROM teaching_groups g
            LEFT JOIN quotas q ON q.group_id = g.id
            WHERE g.active = 1
            GROUP BY g.id, g.name, g.total_capacity
            HAVING COALESCE(SUM(q.capacity), 0) > g.total_capacity
            ORDER BY g.sort_order, g.id
            """
        ).fetchall():
            blockers.append(
                f"{row['name']}各专业配额合计 {row['quota_total']}，"
                f"超过教学组总容量 {row['total_capacity']}"
            )

        for row in connection.execute(
            """
            SELECT m.id, m.name,
                   COUNT(s.id) AS student_count,
                   COALESCE((
                       SELECT SUM(q.capacity)
                       FROM quotas q JOIN teaching_groups g ON g.id = q.group_id
                       WHERE q.major_id = m.id AND g.active = 1
                   ), 0) AS quota_total
            FROM majors m
            JOIN students s ON s.major_id = m.id AND s.active = 1
            WHERE m.active = 1
            GROUP BY m.id, m.name
            ORDER BY m.sort_order, m.id
            """
        ).fetchall():
            if int(row["quota_total"]) < int(row["student_count"]):
                blockers.append(
                    f"{row['name']}配额合计 {row['quota_total']}，少于学生 {row['student_count']} 人"
                )

        if not settings["public_base_url"]:
            blockers.append("尚未设置学生端访问地址")
        elif not settings["public_base_url"].startswith("https://"):
            warnings.append("学生端访问地址不是 HTTPS，仅适合隔离测试或受控校内网络")

        return {"ready": not blockers, "blockers": blockers, "warnings": warnings}

    def activity_list(connection: sqlite3.Connection) -> list[dict[str, Any]]:
        settings = connection.execute(
            "SELECT current_activity_id FROM settings WHERE id = 1"
        ).fetchone()
        current_id = int(settings["current_activity_id"])
        result: list[dict[str, Any]] = []
        for row in connection.execute(
            """
            SELECT id, code, title, status, created_at, opened_at, closed_at,
                   archived_at, summary_json, snapshot_sha256
            FROM activities ORDER BY id DESC
            """
        ).fetchall():
            summary = json.loads(row["summary_json"]) if row["summary_json"] else None
            if row["id"] == current_id:
                totals = connection.execute(
                    """
                    SELECT
                        (SELECT COUNT(*) FROM students WHERE active = 1) AS students,
                        (SELECT COUNT(*) FROM selections se JOIN students s ON s.id = se.student_id
                         WHERE se.revoked_at IS NULL AND s.active = 1) AS selected
                    """
                ).fetchone()
                summary = {
                    "students": int(totals["students"]),
                    "selected": int(totals["selected"]),
                    "unselected": int(totals["students"] - totals["selected"]),
                }
            result.append(
                {
                    "id": row["id"],
                    "code": row["code"],
                    "title": row["title"],
                    "status": row["status"],
                    "current": row["id"] == current_id,
                    "created_at": row["created_at"],
                    "opened_at": row["opened_at"],
                    "closed_at": row["closed_at"],
                    "archived_at": row["archived_at"],
                    "summary": summary,
                    "snapshot_sha256": row["snapshot_sha256"],
                }
            )
        return result

    def student_payload_from_connection(
        connection: sqlite3.Connection, student_id: int
    ) -> dict[str, Any]:
        student = connection.execute(
            """
            SELECT s.id, s.student_no, s.name, s.active, s.major_id,
                   m.name AS major_name, m.active AS major_active
            FROM students s JOIN majors m ON m.id = s.major_id
            WHERE s.id = ?
            """,
            (student_id,),
        ).fetchone()
        if not student or not student["active"]:
            raise HTTPException(status_code=403, detail="学生账号已停用")
        selection = connection.execute(
            """
            SELECT se.group_id, se.selected_at, g.name AS group_name
            FROM selections se
            JOIN teaching_groups g ON g.id = se.group_id
            WHERE se.student_id = ? AND se.revoked_at IS NULL
            """,
            (student_id,),
        ).fetchone()
        groups = connection.execute(
            """
            SELECT g.id, g.name, g.total_capacity,
                   q.capacity,
                   (SELECT COUNT(*) FROM selections sx
                    JOIN students stx ON stx.id = sx.student_id
                    WHERE sx.group_id = g.id AND stx.major_id = ?
                      AND sx.revoked_at IS NULL) AS major_selected,
                   (SELECT COUNT(*) FROM selections sx
                    WHERE sx.group_id = g.id AND sx.revoked_at IS NULL) AS total_selected
            FROM teaching_groups g
            JOIN quotas q ON q.group_id = g.id AND q.major_id = ?
            WHERE g.active = 1
            ORDER BY g.sort_order, g.id
            """,
            (student["major_id"], student["major_id"]),
        ).fetchall()
        settings = setting_dict(connection)
        return {
            "server_now": settings["server_now"],
            "selection_opens_at": settings["selection_opens_at"],
            "phase": settings["phase"],
            "student_login_allowed": settings["student_login_allowed"],
            "status_message": settings["status_message"],
            "student": {
                "id": student["id"],
                "student_no": student["student_no"],
                "name": student["name"],
                "major_id": student["major_id"],
                "major_name": student["major_name"],
            },
            "selection": dict(selection) if selection else None,
            "groups": [
                {
                    "id": row["id"],
                    "name": row["name"],
                    "capacity": row["capacity"],
                    "selected": row["major_selected"],
                    "remaining": max(
                        0,
                        min(
                            row["capacity"] - row["major_selected"],
                            row["total_capacity"] - row["total_selected"],
                        ),
                    ),
                    "full": row["major_selected"] >= row["capacity"]
                    or row["total_selected"] >= row["total_capacity"],
                }
                for row in groups
            ],
            "settings": settings,
        }

    def choose_group(
        *,
        request: Request,
        identity: Identity,
        student_id: int,
        group_id: int,
        source: Literal["student", "admin"],
        operator: str,
        require_open: bool,
    ) -> dict[str, Any]:
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            activity = require_expected_activity(
                request, connection, identity=identity
            )
            if identity.role != source:
                raise HTTPException(status_code=403, detail="登录身份与操作来源不一致")
            if source == "student" and identity.subject_id != student_id:
                raise HTTPException(status_code=403, detail="不能代替其他学生提交选择")
            if operator != str(identity.subject_id):
                raise HTTPException(status_code=403, detail="操作人身份校验失败")
            if require_open:
                phase = activity_phase(activity)
                if phase == "countdown":
                    raise HTTPException(status_code=409, detail="统一倒计时尚未结束，请等待开抢")
                if phase != "open":
                    raise HTTPException(status_code=409, detail="抢选尚未开放或已经结束")
            student = connection.execute(
                """
                SELECT s.id, s.major_id, s.active, m.active AS major_active
                FROM students s JOIN majors m ON m.id = s.major_id
                WHERE s.id = ?
                """,
                (student_id,),
            ).fetchone()
            if not student or not student["active"] or not student["major_active"]:
                raise HTTPException(status_code=403, detail="学生或所属专业已停用")
            existing = connection.execute(
                "SELECT 1 FROM selections WHERE student_id = ? AND revoked_at IS NULL",
                (student_id,),
            ).fetchone()
            if existing:
                raise HTTPException(status_code=409, detail="你已经完成选择，不能重复提交")
            group = connection.execute(
                "SELECT id, active, total_capacity FROM teaching_groups WHERE id = ?",
                (group_id,),
            ).fetchone()
            if not group or not group["active"]:
                raise HTTPException(status_code=404, detail="教学组不存在或已停用")
            quota = connection.execute(
                "SELECT capacity FROM quotas WHERE major_id = ? AND group_id = ?",
                (student["major_id"], group_id),
            ).fetchone()
            if not quota:
                raise HTTPException(status_code=409, detail="该专业尚未配置此教学组配额")
            major_selected = connection.execute(
                """
                SELECT COUNT(*) AS count
                FROM selections se JOIN students s ON s.id = se.student_id
                WHERE se.group_id = ? AND s.major_id = ? AND se.revoked_at IS NULL
                """,
                (group_id, student["major_id"]),
            ).fetchone()["count"]
            total_selected = connection.execute(
                """
                SELECT COUNT(*) AS count FROM selections
                WHERE group_id = ? AND revoked_at IS NULL
                """,
                (group_id,),
            ).fetchone()["count"]
            if major_selected >= quota["capacity"] or total_selected >= group["total_capacity"]:
                raise HTTPException(status_code=409, detail="该教学组名额刚刚已满，请选择其他教学组")
            now = utc_now()
            selection = connection.execute(
                """
                INSERT INTO selections
                    (student_id, group_id, selected_at, source, operator)
                VALUES (?, ?, ?, ?, ?)
                """,
                (student_id, group_id, now, source, operator),
            )
            audit(
                connection,
                actor_type=source,
                actor_id=operator,
                action="selection.create",
                entity_type="selection",
                entity_id=selection.lastrowid,
                details={"student_id": student_id, "group_id": group_id},
            )
            payload = student_payload_from_connection(connection, student_id)
            connection.commit()
            return payload
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    @app.get("/api/health")
    def health() -> dict[str, str]:
        connection = connect(config.database_path)
        try:
            version = int(connection.execute("PRAGMA user_version").fetchone()[0])
            tables = {
                row["name"]
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type = 'table'"
                ).fetchall()
            }
            if version != SCHEMA_VERSION or not {
                "settings",
                "activities",
                "students",
                "selections",
            } <= tables:
                raise RuntimeError("数据库结构未就绪")
            current_activity(connection)
        finally:
            connection.close()
        return {"status": "ok"}

    @app.get("/api/public/info")
    def public_info() -> dict[str, Any]:
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN")
            settings = setting_dict(connection)
            groups = connection.execute(
                """
                SELECT id, name FROM teaching_groups
                WHERE active = 1 ORDER BY sort_order, id
                """
            ).fetchall()
            return {"settings": settings, "group_count": len(groups)}
        finally:
            connection.close()

    @app.get("/api/public/status")
    def public_status() -> dict[str, Any]:
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN")
            settings = setting_dict(connection)
            return {
                "activity_id": settings["activity_id"],
                "status": settings["status"],
                "phase": settings["phase"],
                "server_now": settings["server_now"],
                "selection_opens_at": settings["selection_opens_at"],
                "student_login_allowed": settings["student_login_allowed"],
                "status_message": settings["status_message"],
            }
        finally:
            connection.close()

    @app.post("/api/student/login")
    def student_login(payload: StudentLogin, request: Request, response: Response):
        ip_key = client_key(request, "student-login-ip")
        account_key = principal_key("student-login-account", payload.student_no)
        limiter.check(ip_key, limit=500, window_seconds=300)
        limiter.check(account_key, limit=10, window_seconds=300)
        connection = connect(config.database_path)
        token = new_session_token()
        csrf_token = new_csrf_token()
        now_dt = datetime.now(UTC)
        try:
            connection.execute("BEGIN IMMEDIATE")
            row = connection.execute(
                """
                SELECT id, name, activation_hash, active
                FROM students WHERE student_no = ?
                """,
                (payload.student_no,),
            ).fetchone()
            valid = (
                row
                and row["active"]
                and clean_text(row["name"]) == payload.name
                and verify_activation_code(
                    config.app_secret, payload.activation_code, row["activation_hash"]
                )
            )
            if not valid:
                raise HTTPException(status_code=401, detail="学号、姓名或激活码不正确")
            student_id = int(row["id"])
            student_data = student_payload_from_connection(connection, student_id)
            persist_session(
                connection,
                role="student",
                subject_id=student_id,
                token=token,
                csrf_token=csrf_token,
                now_dt=now_dt,
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        limiter.clear(account_key)
        set_session_cookie(response, "student", token)
        return {"csrf_token": csrf_token, **student_data}

    @app.get("/api/student/me")
    def student_me(request: Request):
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN")
            identity = require_session_from_connection(request, "student", connection)
            payload = student_payload_from_connection(connection, identity.subject_id)
            return {"csrf_token": identity.csrf_token, **payload}
        finally:
            connection.close()

    @app.post("/api/student/heartbeat")
    def student_heartbeat(request: Request):
        """Refresh waiting-room presence without turning every poll into a write.

        The browser may call this every five seconds.  SQLite is updated at most once
        per student per write interval, preserving the write lock for actual selections.
        """

        identity = require_session(request, "student", csrf=True)
        connection = connect(config.database_path)
        try:
            require_expected_activity(request, connection, identity=identity)
            row = connection.execute(
                "SELECT last_seen_at FROM sessions WHERE token_hash = ?",
                (identity.token_hash,),
            ).fetchone()
            now_text = utc_now()
            write_cutoff = (
                parse_utc(now_text) - timedelta(seconds=HEARTBEAT_WRITE_INTERVAL_SECONDS)
            ).isoformat(timespec="seconds")
            if not row or not row["last_seen_at"] or row["last_seen_at"] <= write_cutoff:
                connection.execute("BEGIN IMMEDIATE")
                require_expected_activity(request, connection, identity=identity)
                connection.execute(
                    """
                    UPDATE sessions SET last_seen_at = ?
                    WHERE token_hash = ?
                      AND (last_seen_at IS NULL OR last_seen_at <= ?)
                    """,
                    (now_text, identity.token_hash, write_cutoff),
                )
                connection.commit()
            settings = setting_dict(connection)
            return {
                "ok": True,
                "server_now": settings["server_now"],
                "selection_opens_at": settings["selection_opens_at"],
                "phase": settings["phase"],
                "student_login_allowed": settings["student_login_allowed"],
                "status_message": settings["status_message"],
            }
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    @app.post("/api/student/select")
    def student_select(payload: StudentSelect, request: Request):
        identity = require_session(request, "student", csrf=True)
        result = choose_group(
            request=request,
            identity=identity,
            student_id=identity.subject_id,
            group_id=payload.group_id,
            source="student",
            operator=str(identity.subject_id),
            require_open=True,
        )
        return {"ok": True, **result}

    @app.post("/api/student/logout")
    def student_logout(request: Request, response: Response):
        identity = require_session(request, "student", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("DELETE FROM sessions WHERE token_hash = ?", (identity.token_hash,))
        finally:
            connection.close()
        response.delete_cookie(STUDENT_COOKIE, path="/")
        return {"ok": True}

    @app.post("/api/admin/login")
    def admin_login(payload: AdminLogin, request: Request, response: Response):
        username = payload.username.strip()
        ip_key = client_key(request, "admin-login-ip")
        account_key = principal_key("admin-login-account", username)
        limiter.check(ip_key, limit=50, window_seconds=300)
        limiter.check(account_key, limit=10, window_seconds=300)
        read_connection = connect(config.database_path)
        try:
            candidate = read_connection.execute(
                "SELECT id, password_hash FROM admin_users WHERE username = ?",
                (username,),
            ).fetchone()
        finally:
            read_connection.close()
        if not candidate or not verify_password(payload.password, candidate["password_hash"]):
            raise HTTPException(status_code=401, detail="管理员账号或密码不正确")

        token = new_session_token()
        csrf_token = new_csrf_token()
        now_dt = datetime.now(UTC)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            row = connection.execute(
                "SELECT id, password_hash FROM admin_users WHERE username = ?",
                (username,),
            ).fetchone()
            if not row or not secrets.compare_digest(
                str(row["password_hash"]), str(candidate["password_hash"])
            ):
                raise HTTPException(status_code=401, detail="管理员账号或密码不正确")
            admin_id = int(row["id"])
            persist_session(
                connection,
                role="admin",
                subject_id=admin_id,
                token=token,
                csrf_token=csrf_token,
                now_dt=now_dt,
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        limiter.clear(account_key)
        set_session_cookie(response, "admin", token)
        return {"csrf_token": csrf_token, "username": username}

    @app.get("/api/admin/me")
    def admin_me(request: Request):
        identity = require_session(request, "admin")
        connection = connect(config.database_path)
        try:
            row = connection.execute(
                "SELECT username FROM admin_users WHERE id = ?", (identity.subject_id,)
            ).fetchone()
        finally:
            connection.close()
        if not row:
            raise HTTPException(status_code=401, detail="管理员账号不存在")
        return {"csrf_token": identity.csrf_token, "username": row["username"]}

    @app.post("/api/admin/logout")
    def admin_logout(request: Request, response: Response):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("DELETE FROM sessions WHERE token_hash = ?", (identity.token_hash,))
        finally:
            connection.close()
        response.delete_cookie(ADMIN_COOKIE, path="/")
        return {"ok": True}

    @app.post("/api/admin/password")
    def change_admin_password(payload: PasswordChange, request: Request):
        identity = require_session(request, "admin", csrf=True)
        read_connection = connect(config.database_path)
        try:
            candidate = read_connection.execute(
                "SELECT username, password_hash FROM admin_users WHERE id = ?",
                (identity.subject_id,),
            ).fetchone()
        finally:
            read_connection.close()
        if not candidate or not verify_password(
            payload.current_password, candidate["password_hash"]
        ):
            raise HTTPException(status_code=401, detail="当前密码不正确")
        replacement_hash = hash_password(payload.new_password)

        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            revalidate_identity(request, connection, identity)
            row = connection.execute(
                "SELECT username, password_hash FROM admin_users WHERE id = ?",
                (identity.subject_id,),
            ).fetchone()
            if not row or not secrets.compare_digest(
                str(row["password_hash"]), str(candidate["password_hash"])
            ):
                raise HTTPException(status_code=409, detail="密码已变化，请重新登录后重试")
            connection.execute(
                "UPDATE admin_users SET password_hash = ?, updated_at = ? WHERE id = ?",
                (replacement_hash, utc_now(), identity.subject_id),
            )
            connection.execute(
                """
                DELETE FROM sessions
                WHERE role = 'admin' AND subject_id = ? AND token_hash <> ?
                """,
                (identity.subject_id, identity.token_hash),
            )
            audit(
                connection,
                actor_type="admin",
                actor_id=row["username"],
                action="admin.password.change",
                entity_type="admin_user",
                entity_id=identity.subject_id,
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {"ok": True}

    @app.get("/api/admin/dashboard")
    def admin_dashboard(request: Request):
        require_session(request, "admin")
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN")
            settings = setting_dict(connection)
            presence_cutoff = (
                parse_utc(settings["server_now"])
                - timedelta(seconds=PRESENCE_FRESH_SECONDS)
            ).isoformat(timespec="seconds")
            majors = connection.execute(
                """
                SELECT m.*,
                       (SELECT COUNT(*) FROM students s
                        WHERE s.major_id = m.id AND s.active = 1) AS student_count,
                       (SELECT COUNT(*) FROM selections se
                        JOIN students s ON s.id = se.student_id
                        WHERE s.major_id = m.id AND s.active = 1
                          AND se.revoked_at IS NULL) AS selected_count
                FROM majors m ORDER BY m.sort_order, m.id
                """
            ).fetchall()
            groups = connection.execute(
                """
                SELECT g.*,
                       (SELECT COUNT(*) FROM selections se
                        WHERE se.group_id = g.id AND se.revoked_at IS NULL) AS selected_count
                FROM teaching_groups g ORDER BY g.sort_order, g.id
                """
            ).fetchall()
            quotas = connection.execute(
                """
                SELECT q.major_id, q.group_id, q.capacity,
                       (SELECT COUNT(*) FROM selections se
                        JOIN students s ON s.id = se.student_id
                        WHERE s.major_id = q.major_id AND se.group_id = q.group_id
                          AND se.revoked_at IS NULL) AS selected_count
                FROM quotas q
                """
            ).fetchall()
            totals = connection.execute(
                """
                SELECT
                    (SELECT COUNT(*) FROM students WHERE active = 1) AS students,
                    (SELECT COUNT(*) FROM selections se JOIN students s ON s.id = se.student_id
                     WHERE se.revoked_at IS NULL AND s.active = 1) AS selected
                """
            ).fetchone()
            unselected = connection.execute(
                """
                SELECT s.id, s.student_no, s.name, m.name AS major_name
                FROM students s
                JOIN majors m ON m.id = s.major_id
                LEFT JOIN selections se ON se.student_id = s.id AND se.revoked_at IS NULL
                WHERE s.active = 1 AND se.id IS NULL
                ORDER BY m.sort_order, s.student_no
                """
            ).fetchall()
            recent = connection.execute(
                """
                SELECT s.id AS student_id, s.student_no, s.name, m.name AS major_name,
                       g.name AS group_name, se.selected_at, se.source
                FROM selections se
                JOIN students s ON s.id = se.student_id
                JOIN majors m ON m.id = s.major_id
                JOIN teaching_groups g ON g.id = se.group_id
                WHERE se.revoked_at IS NULL
                ORDER BY se.selected_at DESC LIMIT 20
                """
            ).fetchall()
            students = connection.execute(
                """
                SELECT s.id, s.student_no, s.name, m.name AS major_name, s.active,
                       g.name AS group_name, se.selected_at,
                       CASE WHEN EXISTS (
                           SELECT 1 FROM sessions ss
                           WHERE ss.role = 'student' AND ss.subject_id = s.id
                             AND ss.expires_at > ? AND ss.last_seen_at >= ?
                       ) THEN 1 ELSE 0 END AS entered
                FROM students s
                JOIN majors m ON m.id = s.major_id
                LEFT JOIN selections se
                    ON se.student_id = s.id AND se.revoked_at IS NULL
                LEFT JOIN teaching_groups g ON g.id = se.group_id
                ORDER BY s.active DESC, m.sort_order, s.student_no
                """,
                (session_utc_now(), presence_cutoff),
            ).fetchall()
            presence_rows = connection.execute(
                """
                SELECT s.id, s.student_no, s.name, m.name AS major_name,
                       MAX(ss.last_seen_at) AS last_seen_at
                FROM students s
                JOIN majors m ON m.id = s.major_id
                LEFT JOIN sessions ss
                  ON ss.role = 'student' AND ss.subject_id = s.id
                 AND ss.expires_at > ? AND ss.last_seen_at >= ?
                WHERE s.active = 1
                GROUP BY s.id, s.student_no, s.name, m.name, m.sort_order
                ORDER BY m.sort_order, s.student_no
                """,
                (session_utc_now(), presence_cutoff),
            ).fetchall()
            online_students = [dict(row) for row in presence_rows if row["last_seen_at"]]
            absent_students = [dict(row) for row in presence_rows if not row["last_seen_at"]]
            return {
                "server_now": settings["server_now"],
                "selection_opens_at": settings["selection_opens_at"],
                "phase": settings["phase"],
                "settings": settings,
                "totals": {
                    "students": totals["students"],
                    "selected": totals["selected"],
                    "unselected": totals["students"] - totals["selected"],
                },
                "majors": [dict(row) for row in majors],
                "groups": [dict(row) for row in groups],
                "quotas": [dict(row) for row in quotas],
                "unselected_students": [dict(row) for row in unselected],
                "recent_selections": [dict(row) for row in recent],
                "students": [dict(row) for row in students],
                "presence": {
                    "total": len(presence_rows),
                    "online_count": len(online_students),
                    "absent_count": len(absent_students),
                    "online_students": online_students,
                    "absent_students": absent_students,
                },
                "entered_students": online_students,
                "absent_students": absent_students,
                "readiness": activity_readiness(connection),
                "activities": activity_list(connection),
            }
        finally:
            connection.close()

    @app.patch("/api/admin/settings")
    def update_settings(payload: SettingsUpdate, request: Request):
        identity = require_session(request, "admin", csrf=True)
        values = payload.model_dump(exclude_none=True)
        if not values:
            raise HTTPException(status_code=400, detail="没有需要保存的设置")
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            before = setting_dict(connection)
            title = values.pop("activity_title", None)
            if title is not None:
                connection.execute(
                    "UPDATE activities SET title = ? WHERE id = ?",
                    (title, before["activity_id"]),
                )
                connection.execute(
                    "UPDATE settings SET activity_title = ? WHERE id = 1",
                    (title,),
                )
            settings_updates = {
                "public_base_url": "UPDATE settings SET public_base_url = ?, updated_at = ? WHERE id = 1",
            }
            for column, value in values.items():
                connection.execute(settings_updates[column], (value, utc_now()))
            if not values:
                connection.execute("UPDATE settings SET updated_at = ? WHERE id = 1", (utc_now(),))
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="settings.update",
                entity_type="settings",
                entity_id=1,
                details={"before": before, "changed": payload.model_dump(exclude_none=True)},
            )
            connection.commit()
            return setting_dict(connection)
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    @app.post("/api/admin/status")
    def update_status(payload: StatusUpdate, request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            activity = require_expected_activity(
                request, connection, identity=identity
            )
            old = activity["status"]
            if payload.status == "open" and old != "open":
                readiness = activity_readiness(connection)
                if not readiness["ready"]:
                    raise HTTPException(
                        status_code=409,
                        detail="开放前检查未通过：" + "；".join(readiness["blockers"]),
                    )
            now = utc_now()
            connection.execute(
                """
                UPDATE activities SET status = ?,
                    opened_at = CASE
                        WHEN ? = 'open' THEN COALESCE(opened_at, ?)
                        WHEN selection_opens_at IS NOT NULL
                             AND selection_opens_at > ?
                             AND opened_at = selection_opens_at THEN NULL
                        ELSE opened_at
                    END,
                    closed_at = CASE WHEN ? = 'closed' THEN ? ELSE closed_at END,
                    selection_opens_at = CASE WHEN ? = 'open' THEN ? ELSE NULL END
                WHERE id = ?
                """,
                (
                    payload.status,
                    payload.status,
                    now,
                    now,
                    payload.status,
                    now,
                    payload.status,
                    now,
                    activity["id"],
                ),
            )
            connection.execute(
                "UPDATE settings SET status = ?, updated_at = ? WHERE id = 1",
                (payload.status, now),
            )
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action=f"activity.{payload.status}",
                entity_type="settings",
                entity_id=1,
                details={"from": old, "to": payload.status},
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {"status": payload.status}

    @app.post("/api/admin/countdown")
    @app.post("/api/admin/start-countdown")
    def start_countdown(request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            activity = require_expected_activity(request, connection, identity=identity)
            if activity["status"] == "open":
                raise HTTPException(status_code=409, detail="本场抢选已在倒计时或已开放")
            readiness = activity_readiness(connection)
            if not readiness["ready"]:
                raise HTTPException(
                    status_code=409,
                    detail="开始前检查未通过：" + "；".join(readiness["blockers"]),
                )
            now = utc_now()
            target_timespec = "milliseconds" if "." in now else "seconds"
            opens_at = (
                parse_utc(now) + timedelta(seconds=COUNTDOWN_SECONDS)
            ).isoformat(timespec=target_timespec)
            connection.execute(
                """
                UPDATE activities
                SET status = 'open', selection_opens_at = ?,
                    opened_at = ?
                WHERE id = ? AND status = 'closed'
                """,
                (opens_at, opens_at, activity["id"]),
            )
            connection.execute(
                "UPDATE settings SET status = 'open', updated_at = ? WHERE id = 1",
                (now,),
            )
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="activity.countdown.start",
                entity_type="activity",
                entity_id=activity["id"],
                details={
                    "countdown_seconds": COUNTDOWN_SECONDS,
                    "selection_opens_at": opens_at,
                },
            )
            connection.commit()
            settings = setting_dict(connection)
            return {
                "status": settings["status"],
                "phase": settings["phase"],
                "server_now": settings["server_now"],
                "selection_opens_at": settings["selection_opens_at"],
            }
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    @app.get("/api/admin/activities")
    def list_activities(request: Request):
        require_session(request, "admin")
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN")
            return activity_list(connection)
        finally:
            connection.close()

    @app.post("/api/admin/activities")
    def create_activity(payload: ActivityCreate, request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            previous = current_activity(connection)
            require_expected_activity(request, connection, identity=identity)
            if int(previous["id"]) != payload.previous_activity_id:
                raise HTTPException(
                    status_code=409,
                    detail="当前活动已经变化，请刷新管理端后重试",
                )
            if previous["status"] == "open":
                raise HTTPException(status_code=409, detail="请先关闭当前抢选再新建活动")

            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="activity.archive",
                entity_type="activity",
                entity_id=previous["id"],
                details={
                    "next_title": payload.title,
                    "copy_structure": payload.copy_structure,
                },
                activity_id=int(previous["id"]),
            )
            now = utc_now()
            snapshot, snapshot_hash = activity_snapshot(
                connection, int(previous["id"]), archived_at=now
            )
            snapshot_json = json.dumps(
                snapshot,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            )
            active_student_ids = {
                int(row["id"]) for row in snapshot["students"] if row["active"]
            }
            totals = {
                "students": len(active_student_ids),
                "selected": sum(
                    1
                    for row in snapshot["selections"]
                    if row["revoked_at"] is None
                    and int(row["student_id"]) in active_student_ids
                ),
            }
            totals["unselected"] = totals["students"] - totals["selected"]
            connection.execute(
                """
                UPDATE activities
                SET status = 'archived', archived_at = ?, summary_json = ?,
                    snapshot_json = ?, snapshot_sha256 = ?
                WHERE id = ?
                """,
                (
                    now,
                    json.dumps(totals, ensure_ascii=False, separators=(",", ":")),
                    snapshot_json,
                    snapshot_hash,
                    previous["id"],
                ),
            )

            code = payload.code or (
                f"activity-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}-{secrets.token_hex(2)}"
            )
            cursor = connection.execute(
                """
                INSERT INTO activities (code, title, status, created_at, closed_at)
                VALUES (?, ?, 'closed', ?, ?)
                """,
                (code, payload.title, now, now),
            )
            activity_id = int(cursor.lastrowid)

            previous_majors = connection.execute(
                "SELECT * FROM majors ORDER BY sort_order, id"
            ).fetchall()
            previous_groups = connection.execute(
                "SELECT * FROM teaching_groups ORDER BY sort_order, id"
            ).fetchall()
            previous_quotas = connection.execute("SELECT * FROM quotas").fetchall()

            connection.execute("DELETE FROM sessions WHERE role = 'student'")
            connection.execute("DELETE FROM selections")
            connection.execute("DELETE FROM students")
            connection.execute("DELETE FROM quotas")
            connection.execute("DELETE FROM teaching_groups")
            connection.execute("DELETE FROM majors")

            if payload.copy_structure:
                major_ids: dict[int, int] = {}
                for row in previous_majors:
                    created = connection.execute(
                        """
                        INSERT INTO majors (code, name, active, sort_order, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (row["code"], row["name"], row["active"], row["sort_order"], now, now),
                    )
                    major_ids[int(row["id"])] = int(created.lastrowid)
                group_ids: dict[int, int] = {}
                for row in previous_groups:
                    created = connection.execute(
                        """
                        INSERT INTO teaching_groups
                            (code, name, total_capacity, active, sort_order, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            row["code"], row["name"], row["total_capacity"], row["active"],
                            row["sort_order"], now, now,
                        ),
                    )
                    group_ids[int(row["id"])] = int(created.lastrowid)
                for row in previous_quotas:
                    connection.execute(
                        """
                        INSERT INTO quotas (major_id, group_id, capacity, updated_at)
                        VALUES (?, ?, ?, ?)
                        """,
                        (
                            major_ids[int(row["major_id"])],
                            group_ids[int(row["group_id"])],
                            row["capacity"],
                            now,
                        ),
                    )

            connection.execute(
                """
                UPDATE settings SET current_activity_id = ?, activity_title = ?,
                    status = 'closed', updated_at = ? WHERE id = 1
                """,
                (activity_id, payload.title, now),
            )
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="activity.create",
                entity_type="activity",
                entity_id=activity_id,
                details={
                    "code": code,
                    "title": payload.title,
                    "copy_structure": payload.copy_structure,
                    "archived_activity_id": previous["id"],
                    "archived_snapshot_sha256": snapshot_hash,
                },
                activity_id=activity_id,
            )
            connection.commit()
            return setting_dict(connection)
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    @app.get("/api/admin/activities/{activity_id}/archive.json")
    def export_activity_archive(activity_id: int, request: Request):
        require_session(request, "admin")
        connection = connect(config.database_path)
        try:
            row = connection.execute(
                """
                SELECT id, code, title, status, created_at, opened_at, closed_at,
                       archived_at, summary_json, snapshot_json, snapshot_sha256
                FROM activities WHERE id = ?
                """,
                (activity_id,),
            ).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="活动不存在")
            if row["status"] != "archived":
                raise HTTPException(status_code=409, detail="当前活动尚未归档")
            if not row["archived_at"] or not row["snapshot_json"]:
                raise HTTPException(
                    status_code=500, detail="归档校验失败，请先运行数据库检查"
                )
            digest = hashlib.sha256(row["snapshot_json"].encode("utf-8")).hexdigest()
            if not row["snapshot_sha256"] or not hmac.compare_digest(
                digest, str(row["snapshot_sha256"])
            ):
                raise HTTPException(status_code=500, detail="归档校验失败，请先运行数据库检查")
            try:
                summary = json.loads(row["summary_json"])
                snapshot = json.loads(row["snapshot_json"])
                archived_activity = snapshot["activity"]
                archived_at = datetime.fromisoformat(str(row["archived_at"]))
            except (json.JSONDecodeError, KeyError, TypeError):
                raise HTTPException(
                    status_code=500, detail="归档校验失败，请先运行数据库检查"
                )
            except ValueError:
                raise HTTPException(
                    status_code=500, detail="归档校验失败，请先运行数据库检查"
                )
            if (
                archived_at.utcoffset() is None
                or snapshot.get("archived_at") != row["archived_at"]
            ):
                raise HTTPException(
                    status_code=500, detail="归档校验失败，请先运行数据库检查"
                )
            expected_activity = {
                "id": row["id"],
                "code": row["code"],
                "title": row["title"],
                "created_at": row["created_at"],
                "opened_at": row["opened_at"],
                "closed_at": row["closed_at"],
            }
            if not isinstance(archived_activity, dict) or any(
                archived_activity.get(key) != value
                for key, value in expected_activity.items()
            ):
                raise HTTPException(
                    status_code=500, detail="归档校验失败，请先运行数据库检查"
                )
            active_student_ids = {
                int(student["id"])
                for student in snapshot["students"]
                if student.get("active")
            }
            expected_summary = {
                "students": len(active_student_ids),
                "selected": sum(
                    1
                    for selection in snapshot["selections"]
                    if selection.get("revoked_at") is None
                    and int(selection["student_id"]) in active_student_ids
                ),
            }
            expected_summary["unselected"] = (
                expected_summary["students"] - expected_summary["selected"]
            )
            if summary != expected_summary:
                raise HTTPException(
                    status_code=500, detail="归档校验失败，请先运行数据库检查"
                )
            return Response(
                content=row["snapshot_json"].encode("utf-8"),
                media_type="application/json",
                headers={
                    "Content-Disposition": f'attachment; filename="{row["code"]}-archive.json"',
                    "X-Archive-SHA256": row["snapshot_sha256"],
                    "Cache-Control": "no-store",
                },
            )
        finally:
            connection.close()

    @app.delete("/api/admin/activities/{activity_id}")
    def delete_activity_archive(
        activity_id: int, payload: ArchiveDelete, request: Request
    ):
        """Permanently remove one verified archive, never the live activity.

        The fixed confirmation literal is an independent server-side guard; the
        UI presents two human confirmations before sending it.  Only summary
        metadata is retained in the current activity's audit trail.
        """

        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            current = require_expected_activity(
                request, connection, identity=identity
            )
            row = connection.execute(
                """
                SELECT id, code, status, archived_at, summary_json,
                       snapshot_json, snapshot_sha256
                FROM activities WHERE id = ?
                """,
                (activity_id,),
            ).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="活动不存在")
            if int(row["id"]) == int(current["id"]) or row["status"] != "archived":
                raise HTTPException(status_code=409, detail="只能删除已归档的历史活动")
            if not row["snapshot_json"] or not row["snapshot_sha256"]:
                raise HTTPException(
                    status_code=500, detail="归档校验失败，请先运行数据库检查"
                )
            digest = hashlib.sha256(row["snapshot_json"].encode("utf-8")).hexdigest()
            if not hmac.compare_digest(digest, str(row["snapshot_sha256"])):
                raise HTTPException(
                    status_code=500, detail="归档校验失败，请先运行数据库检查"
                )
            try:
                summary = json.loads(row["summary_json"])
            except (json.JSONDecodeError, TypeError) as exc:
                raise HTTPException(
                    status_code=500, detail="归档校验失败，请先运行数据库检查"
                ) from exc
            safe_summary = {
                key: int(summary[key]) for key in ("students", "selected", "unselected")
            }

            # Archived audit rows are part of the immutable snapshot being
            # deleted.  Record a non-PII tombstone under the live activity first,
            # then remove rows that would otherwise RESTRICT the archive delete.
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="activity.archive.delete",
                entity_type="activity_archive",
                entity_id=activity_id,
                details={
                    "code": row["code"],
                    "archived_at": row["archived_at"],
                    "summary": safe_summary,
                    "snapshot_sha256": row["snapshot_sha256"],
                },
                activity_id=int(current["id"]),
            )
            connection.execute(
                "DELETE FROM audit_logs WHERE activity_id = ?", (activity_id,)
            )
            connection.execute("DELETE FROM activities WHERE id = ?", (activity_id,))
            connection.commit()
            return {"ok": True, "deleted_activity_id": activity_id}
        except KeyError as exc:
            connection.rollback()
            raise HTTPException(
                status_code=500, detail="归档校验失败，请先运行数据库检查"
            ) from exc
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    @app.post("/api/admin/majors")
    def create_major(payload: MajorCreate, request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            ensure_closed(connection)
            max_sort = connection.execute(
                "SELECT COALESCE(MAX(sort_order), 0) AS value FROM majors"
            ).fetchone()["value"]
            now = utc_now()
            cursor = connection.execute(
                """
                INSERT INTO majors (code, name, active, sort_order, created_at, updated_at)
                VALUES (?, ?, 1, ?, ?, ?)
                """,
                (f"major-{secrets.token_hex(4)}", payload.name, max_sort + 10, now, now),
            )
            major_id = int(cursor.lastrowid)
            connection.execute(
                """
                INSERT INTO quotas (major_id, group_id, capacity, updated_at)
                SELECT ?, id, 0, ? FROM teaching_groups
                """,
                (major_id, now),
            )
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="major.create",
                entity_type="major",
                entity_id=major_id,
                details={"name": payload.name},
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {"id": major_id, "name": payload.name}

    @app.patch("/api/admin/majors/{major_id}")
    def update_major(major_id: int, payload: MajorUpdate, request: Request):
        identity = require_session(request, "admin", csrf=True)
        values = payload.model_dump(exclude_none=True)
        if "active" in values:
            values["active"] = int(values["active"])
        if not values:
            raise HTTPException(status_code=400, detail="没有需要保存的内容")
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            ensure_closed(connection)
            before = connection.execute("SELECT * FROM majors WHERE id = ?", (major_id,)).fetchone()
            if not before:
                raise HTTPException(status_code=404, detail="专业不存在")
            if values.get("active") == 0 and bool(before["active"]):
                active_students = int(
                    connection.execute(
                        "SELECT COUNT(*) FROM students WHERE major_id = ? AND active = 1",
                        (major_id,),
                    ).fetchone()[0]
                )
                if active_students:
                    raise HTTPException(
                        status_code=409,
                        detail=(
                            f"该专业仍有 {active_students} 名有效学生，不能停用；"
                            "请先通过学生名单将其迁移或停用"
                        ),
                    )
            major_updates = {
                "name": "UPDATE majors SET name = ?, updated_at = ? WHERE id = ?",
                "active": "UPDATE majors SET active = ?, updated_at = ? WHERE id = ?",
                "sort_order": "UPDATE majors SET sort_order = ?, updated_at = ? WHERE id = ?",
            }
            for column, value in values.items():
                connection.execute(major_updates[column], (value, utc_now(), major_id))
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="major.update",
                entity_type="major",
                entity_id=major_id,
                details={"before": dict(before), "changed": values},
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {"ok": True}

    @app.delete("/api/admin/majors/{major_id}")
    def delete_major(major_id: int, request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            ensure_closed(connection)
            row = connection.execute("SELECT name FROM majors WHERE id = ?", (major_id,)).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="专业不存在")
            students = connection.execute(
                "SELECT COUNT(*) AS count FROM students WHERE major_id = ?", (major_id,)
            ).fetchone()["count"]
            if students:
                raise HTTPException(status_code=409, detail="该专业已有学生，不能删除；可改为停用")
            connection.execute("DELETE FROM majors WHERE id = ?", (major_id,))
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="major.delete",
                entity_type="major",
                entity_id=major_id,
                details={"name": row["name"]},
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {"ok": True}

    @app.post("/api/admin/groups")
    def create_group(payload: GroupCreate, request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            ensure_closed(connection)
            max_sort = connection.execute(
                "SELECT COALESCE(MAX(sort_order), 0) AS value FROM teaching_groups"
            ).fetchone()["value"]
            now = utc_now()
            cursor = connection.execute(
                """
                INSERT INTO teaching_groups
                    (code, name, total_capacity, active, sort_order, created_at, updated_at)
                VALUES (?, ?, ?, 1, ?, ?, ?)
                """,
                (
                    f"group-{secrets.token_hex(4)}",
                    payload.name,
                    payload.total_capacity,
                    max_sort + 10,
                    now,
                    now,
                ),
            )
            group_id = int(cursor.lastrowid)
            connection.execute(
                """
                INSERT INTO quotas (major_id, group_id, capacity, updated_at)
                SELECT id, ?, 0, ? FROM majors
                """,
                (group_id, now),
            )
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="group.create",
                entity_type="teaching_group",
                entity_id=group_id,
                details={"name": payload.name, "total_capacity": payload.total_capacity},
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {"id": group_id, "name": payload.name}

    @app.patch("/api/admin/groups/{group_id}")
    def update_group(group_id: int, payload: GroupUpdate, request: Request):
        identity = require_session(request, "admin", csrf=True)
        values = payload.model_dump(exclude_none=True)
        if "active" in values:
            values["active"] = int(values["active"])
        if not values:
            raise HTTPException(status_code=400, detail="没有需要保存的内容")
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            ensure_closed(connection)
            before = connection.execute(
                "SELECT * FROM teaching_groups WHERE id = ?", (group_id,)
            ).fetchone()
            if not before:
                raise HTTPException(status_code=404, detail="教学组不存在")
            selected = connection.execute(
                "SELECT COUNT(*) AS count FROM selections WHERE group_id = ? AND revoked_at IS NULL",
                (group_id,),
            ).fetchone()["count"]
            if "total_capacity" in values and values["total_capacity"] < selected:
                raise HTTPException(
                    status_code=409,
                    detail=f"总容量不能小于当前已选人数 {selected}",
                )
            if values.get("active") == 0 and bool(before["active"]) and selected:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"该教学组仍有 {selected} 条有效选择，不能停用；"
                        "请先撤销这些选择"
                    ),
                )
            quota_adjustments: list[dict[str, Any]] = []
            capacity_changed = (
                "total_capacity" in values
                and int(values["total_capacity"]) != int(before["total_capacity"])
            )
            if capacity_changed:
                quota_adjustments = rebalance_group_quotas(
                    connection, group_id, int(values["total_capacity"])
                )
            group_updates = {
                "name": "UPDATE teaching_groups SET name = ?, updated_at = ? WHERE id = ?",
                "active": "UPDATE teaching_groups SET active = ?, updated_at = ? WHERE id = ?",
                "sort_order": (
                    "UPDATE teaching_groups SET sort_order = ?, updated_at = ? WHERE id = ?"
                ),
            }
            for column, value in values.items():
                if column == "total_capacity":
                    continue
                connection.execute(group_updates[column], (value, utc_now(), group_id))
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="group.update",
                entity_type="teaching_group",
                entity_id=group_id,
                details={
                    "before": dict(before),
                    "changed": values,
                    "quota_adjustments": quota_adjustments,
                },
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {
            "ok": True,
            "quotas_adjusted": bool(quota_adjustments),
            "quota_adjustments": quota_adjustments,
        }

    @app.delete("/api/admin/groups/{group_id}")
    def delete_group(group_id: int, request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            ensure_closed(connection)
            row = connection.execute(
                "SELECT name FROM teaching_groups WHERE id = ?", (group_id,)
            ).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="教学组不存在")
            selections = connection.execute(
                "SELECT COUNT(*) AS count FROM selections WHERE group_id = ?", (group_id,)
            ).fetchone()["count"]
            if selections:
                raise HTTPException(status_code=409, detail="该教学组已有历史选择，不能删除；可改为停用")
            connection.execute("DELETE FROM teaching_groups WHERE id = ?", (group_id,))
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="group.delete",
                entity_type="teaching_group",
                entity_id=group_id,
                details={"name": row["name"]},
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {"ok": True}

    @app.put("/api/admin/quotas/{major_id}/{group_id}")
    def update_quota(major_id: int, group_id: int, payload: QuotaUpdate, request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            ensure_closed(connection)
            current = connection.execute(
                "SELECT capacity FROM quotas WHERE major_id = ? AND group_id = ?",
                (major_id, group_id),
            ).fetchone()
            if not current:
                raise HTTPException(status_code=404, detail="配额单元不存在")
            selected = connection.execute(
                """
                SELECT COUNT(*) AS count
                FROM selections se JOIN students s ON s.id = se.student_id
                WHERE s.major_id = ? AND se.group_id = ? AND se.revoked_at IS NULL
                """,
                (major_id, group_id),
            ).fetchone()["count"]
            if payload.capacity < selected:
                raise HTTPException(status_code=409, detail=f"配额不能小于当前已选人数 {selected}")
            other_sum = connection.execute(
                """
                SELECT COALESCE(SUM(capacity), 0) AS total FROM quotas
                WHERE group_id = ? AND major_id <> ?
                """,
                (group_id, major_id),
            ).fetchone()["total"]
            total_capacity = connection.execute(
                "SELECT total_capacity FROM teaching_groups WHERE id = ?", (group_id,)
            ).fetchone()["total_capacity"]
            if other_sum + payload.capacity > total_capacity:
                raise HTTPException(
                    status_code=409,
                    detail=f"各专业配额合计不能超过教学组总容量 {total_capacity}",
                )
            connection.execute(
                """
                UPDATE quotas SET capacity = ?, updated_at = ?
                WHERE major_id = ? AND group_id = ?
                """,
                (payload.capacity, utc_now(), major_id, group_id),
            )
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="quota.update",
                entity_type="quota",
                entity_id=f"{major_id}:{group_id}",
                details={"from": current["capacity"], "to": payload.capacity},
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {"ok": True}

    @app.post("/api/admin/students/import")
    async def import_students(
        request: Request,
        file: Annotated[
            UploadFile | None,
            File(description="兼容旧客户端的单个 CSV、XLS 或 XLSX 名单"),
        ] = None,
        files: Annotated[
            list[UploadFile] | None,
            File(description="可重复提交的 CSV、XLS 或 XLSX 名单集合"),
        ] = None,
        mode: Literal["merge", "sync"] = "merge",
        regenerate_existing: bool = False,
    ):
        identity = require_session(request, "admin", csrf=True)
        uploads = ([file] if file is not None else []) + list(files or [])
        if not uploads:
            raise HTTPException(status_code=400, detail="请至少上传一个名单文件")
        if len(uploads) > MAX_IMPORT_FILES:
            raise HTTPException(
                status_code=413,
                detail=f"一次最多上传 {MAX_IMPORT_FILES} 个名单文件",
            )

        upload_data: list[tuple[str, bytes]] = []
        try:
            for upload in uploads:
                content = await upload.read(MAX_ROSTER_FILE_BYTES + 1)
                if len(content) > MAX_ROSTER_FILE_BYTES:
                    raise HTTPException(status_code=413, detail="单个名单文件不能超过 1 MB")
                upload_data.append((upload.filename or "upload", content))
        finally:
            for upload in uploads:
                await upload.close()
        try:
            rows = parse_roster_files(upload_data)
        except RosterParseError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        finally:
            upload_data.clear()
        if len(rows) > MAX_ROSTER_ROWS:
            raise HTTPException(
                status_code=413,
                detail=f"一次最多导入 {MAX_ROSTER_ROWS} 名学生",
            )

        seen_numbers: set[str] = set()
        for row in rows:
            if row.student_no in seen_numbers:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"第 {row.file_index} 个文件第 {row.line_number} 行的学号"
                        "在本次名单中重复"
                    ),
                )
            seen_numbers.add(row.student_no)
            if (
                len(row.student_no) > 40
                or len(row.name) > 80
                or len(row.major_name) > 80
            ):
                raise HTTPException(
                    status_code=400,
                    detail=f"第 {row.file_index} 个文件第 {row.line_number} 行字段长度超限",
                )

        connection = connect(config.database_path)
        created = 0
        updated = 0
        deactivated = 0
        rotated = 0
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            ensure_closed(connection)
            major_map = {
                row["name"]: int(row["id"])
                for row in connection.execute(
                    "SELECT id, name FROM majors WHERE active = 1"
                ).fetchall()
            }
            for row in rows:
                student_no = row.student_no
                name = row.name
                major_name = row.major_name
                code = row.activation_code
                if major_name not in major_map:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"第 {row.file_index} 个文件第 {row.line_number} 行的专业"
                            f"“{major_name}”不存在或已停用"
                        ),
                    )
                existing = connection.execute(
                    """
                    SELECT id, name, major_id, activation_hash,
                           activation_ciphertext, active
                    FROM students WHERE student_no = ?
                    """,
                    (student_no,),
                ).fetchone()
                if existing:
                    active_selection = connection.execute(
                        "SELECT 1 FROM selections WHERE student_id = ? AND revoked_at IS NULL",
                        (existing["id"],),
                    ).fetchone()
                    if active_selection and existing["major_id"] != major_map[major_name]:
                        raise HTTPException(
                            status_code=409,
                            detail=f"学号 {student_no} 已有选择，不能更改所属专业",
                        )
                    profile_changed = (
                        clean_text(existing["name"]) != name
                        or int(existing["major_id"]) != int(major_map[major_name])
                    )
                    credential_changed = not verify_activation_code(
                        config.app_secret, code, existing["activation_hash"]
                    )
                    access_changed = not bool(existing["active"])
                    connection.execute(
                        """
                        UPDATE students SET name = ?, major_id = ?, activation_hash = ?,
                                            activation_ciphertext = ?,
                                            active = 1, updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            name,
                            major_map[major_name],
                            activation_code_hash(config.app_secret, code),
                            encrypt_activation_code(config.app_secret, student_no, code),
                            utc_now(),
                            existing["id"],
                        ),
                    )
                    if credential_changed:
                        rotated += 1
                    if credential_changed or profile_changed or access_changed:
                        connection.execute(
                            "DELETE FROM sessions WHERE role = 'student' AND subject_id = ?",
                            (existing["id"],),
                        )
                    updated += 1
                else:
                    now = utc_now()
                    connection.execute(
                        """
                        INSERT INTO students
                            (student_no, name, major_id, activation_hash,
                             activation_ciphertext, active, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, 1, ?, ?)
                        """,
                        (
                            student_no,
                            name,
                            major_map[major_name],
                            activation_code_hash(config.app_secret, code),
                            encrypt_activation_code(config.app_secret, student_no, code),
                            now,
                            now,
                        ),
                    )
                    created += 1

            if mode == "sync":
                missing_students = [
                    row
                    for row in connection.execute(
                        "SELECT id, student_no, name FROM students WHERE active = 1"
                    ).fetchall()
                    if row["student_no"] not in seen_numbers
                ]
                blocked = []
                for missing in missing_students:
                    has_selection = connection.execute(
                        "SELECT 1 FROM selections WHERE student_id = ? AND revoked_at IS NULL",
                        (missing["id"],),
                    ).fetchone()
                    if has_selection:
                        blocked.append(f"{missing['student_no']} {missing['name']}")
                if blocked:
                    preview = "、".join(blocked[:5])
                    suffix = "等" if len(blocked) > 5 else ""
                    raise HTTPException(
                        status_code=409,
                        detail=f"同步名单会移除已有选择的学生：{preview}{suffix}；请先撤销其选择",
                    )
                for missing in missing_students:
                    connection.execute(
                        "UPDATE students SET active = 0, updated_at = ? WHERE id = ?",
                        (utc_now(), missing["id"]),
                    )
                    connection.execute(
                        "DELETE FROM sessions WHERE role = 'student' AND subject_id = ?",
                        (missing["id"],),
                    )
                deactivated = len(missing_students)
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="students.import",
                entity_type="student",
                entity_id="batch",
                details={
                    "mode": mode,
                    "created": created,
                    "updated": updated,
                    "deactivated": deactivated,
                    "rotated": rotated,
                    "file_count": len(uploads),
                    "row_count": len(rows),
                    "credential_source": "document_number_suffix",
                },
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {
            "created": created,
            "updated": updated,
            "deactivated": deactivated,
            "rotated": rotated,
            "file_count": len(uploads),
            "row_count": len(rows),
            "activation_code_policy": "normalized_document_number_last_6",
        }

    @app.post("/api/admin/students/{student_id}/activation-code")
    def reset_student_activation_code(student_id: int, request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            student = connection.execute(
                "SELECT id FROM students WHERE id = ?",
                (student_id,),
            ).fetchone()
            if not student:
                raise HTTPException(status_code=404, detail="学生不存在")
            raise HTTPException(
                status_code=409,
                detail=(
                    "激活码只能由证件号规范化后的末 6 位生成；"
                    "请重新导入包含证件号的名单以更新该学生凭据"
                ),
            )
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    @app.post("/api/admin/students/{student_id}/activation-code/reveal")
    def reveal_student_activation_code(student_id: int, request: Request):
        identity = require_session(request, "admin", csrf=True)
        limiter.check(
            principal_key("activation-code-reveal", str(identity.subject_id)),
            limit=500,
            window_seconds=300,
        )
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            student = connection.execute(
                """
                SELECT s.id, s.student_no, s.name, s.activation_hash,
                       s.activation_ciphertext, m.name AS major_name
                FROM students s JOIN majors m ON m.id = s.major_id
                WHERE s.id = ?
                """,
                (student_id,),
            ).fetchone()
            if not student:
                raise HTTPException(status_code=404, detail="学生不存在")
            if not student["activation_ciphertext"]:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "该学生为历史哈希凭据，原激活码无法显示；"
                        "请重新导入包含证件号的名单"
                    ),
                )
            try:
                code = decrypt_activation_code(
                    config.app_secret,
                    student["student_no"],
                    student["activation_ciphertext"],
                )
            except Exception as exc:
                raise HTTPException(
                    status_code=409,
                    detail="激活码密文无法校验，请重新导入包含证件号的名单",
                ) from exc
            if not verify_activation_code(
                config.app_secret, code, student["activation_hash"]
            ):
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "激活码密文与登录凭据不一致；"
                        "请重新导入包含证件号的名单"
                    ),
                )
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="student.activation_code.reveal",
                entity_type="student",
                entity_id=student_id,
                details={"student_no": student["student_no"]},
            )
            connection.commit()
            return JSONResponse(
                content={
                    "credential": {
                        "student_no": student["student_no"],
                        "name": student["name"],
                        "major": student["major_name"],
                        "activation_code": code,
                    }
                },
                headers={
                    "Cache-Control": "no-store, private",
                    "Pragma": "no-cache",
                },
            )
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    @app.post("/api/admin/selections")
    def admin_assign(payload: AdminAssign, request: Request):
        identity = require_session(request, "admin", csrf=True)
        choose_group(
            request=request,
            identity=identity,
            student_id=payload.student_id,
            group_id=payload.group_id,
            source="admin",
            operator=str(identity.subject_id),
            require_open=False,
        )
        return {"ok": True}

    @app.post("/api/admin/selections/revoke")
    def revoke_selection(payload: RevokeSelection, request: Request):
        identity = require_session(request, "admin", csrf=True)
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN IMMEDIATE")
            require_expected_activity(request, connection, identity=identity)
            selection = connection.execute(
                """
                SELECT id, group_id FROM selections
                WHERE student_id = ? AND revoked_at IS NULL
                """,
                (payload.student_id,),
            ).fetchone()
            if not selection:
                raise HTTPException(status_code=404, detail="该学生没有可撤销的当前选择")
            now = utc_now()
            connection.execute(
                "UPDATE selections SET revoked_at = ?, revoked_by = ? WHERE id = ?",
                (now, str(identity.subject_id), selection["id"]),
            )
            audit(
                connection,
                actor_type="admin",
                actor_id=identity.subject_id,
                action="selection.revoke",
                entity_type="selection",
                entity_id=selection["id"],
                details={
                    "student_id": payload.student_id,
                    "group_id": selection["group_id"],
                    "reason": payload.reason,
                },
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {"ok": True}

    def spreadsheet_safe_cell(value: Any) -> str:
        text = "" if value is None else str(value)
        visible = text.lstrip(" \t\r\n")
        if text.startswith(("\t", "\r", "\n")) or visible.startswith(("=", "+", "-", "@")):
            return "'" + text
        return text

    def csv_response(filename: str, headers: list[str], rows: list[list[Any]]):
        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator="\n")
        writer.writerow(headers)
        writer.writerows(
            [[spreadsheet_safe_cell(value) for value in row] for row in rows]
        )
        content = "\ufeff" + buffer.getvalue()
        return StreamingResponse(
            iter([content.encode("utf-8")]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.get("/api/admin/export/selections.csv")
    def export_selections(request: Request, activity_id: int):
        require_session(request, "admin")
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN")
            activity = current_activity(connection)
            if int(activity["id"]) != activity_id:
                raise HTTPException(status_code=409, detail="当前活动已经变化，请刷新页面后重试")
            rows = connection.execute(
                """
                SELECT s.student_no, s.name, m.name AS major_name,
                       g.name AS group_name, se.selected_at, se.source
                FROM selections se
                JOIN students s ON s.id = se.student_id
                JOIN majors m ON m.id = s.major_id
                JOIN teaching_groups g ON g.id = se.group_id
                WHERE se.revoked_at IS NULL
                ORDER BY m.sort_order, g.sort_order, s.student_no
                """
            ).fetchall()
        finally:
            connection.close()
        return csv_response(
            f"{activity['code']}-selections.csv",
            ["学号", "姓名", "专业", "教学组", "选择时间", "来源"],
            [[row[key] for key in row.keys()] for row in rows],
        )

    @app.get("/api/admin/export/unselected.csv")
    def export_unselected(request: Request, activity_id: int):
        require_session(request, "admin")
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN")
            activity = current_activity(connection)
            if int(activity["id"]) != activity_id:
                raise HTTPException(status_code=409, detail="当前活动已经变化，请刷新页面后重试")
            rows = connection.execute(
                """
                SELECT s.student_no, s.name, m.name AS major_name
                FROM students s JOIN majors m ON m.id = s.major_id
                LEFT JOIN selections se ON se.student_id = s.id AND se.revoked_at IS NULL
                WHERE s.active = 1 AND se.id IS NULL
                ORDER BY m.sort_order, s.student_no
                """
            ).fetchall()
        finally:
            connection.close()
        return csv_response(
            f"{activity['code']}-unselected.csv",
            ["学号", "姓名", "专业"],
            [[row["student_no"], row["name"], row["major_name"]] for row in rows],
        )

    @app.get("/api/admin/export/results.csv")
    def export_complete_results(request: Request, activity_id: int):
        require_session(request, "admin")
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN")
            activity = current_activity(connection)
            if int(activity["id"]) != activity_id:
                raise HTTPException(status_code=409, detail="当前活动已经变化，请刷新页面后重试")
            rows = connection.execute(
                """
                SELECT s.student_no, s.name, m.name AS major_name,
                       CASE WHEN se.id IS NULL THEN '未选' ELSE '已选' END AS selection_status,
                       g.name AS group_name, se.selected_at
                FROM students s
                JOIN majors m ON m.id = s.major_id
                LEFT JOIN selections se
                  ON se.student_id = s.id AND se.revoked_at IS NULL
                LEFT JOIN teaching_groups g ON g.id = se.group_id
                WHERE s.active = 1
                ORDER BY m.sort_order, s.student_no
                """
            ).fetchall()
        finally:
            connection.close()
        return csv_response(
            f"{activity['code']}-results.csv",
            ["学号", "姓名", "专业", "选择状态", "教学组", "选择时间"],
            [
                [
                    row["student_no"],
                    row["name"],
                    row["major_name"],
                    row["selection_status"],
                    row["group_name"],
                    row["selected_at"],
                ]
                for row in rows
            ],
        )

    @app.get("/api/admin/export/results.xlsx")
    def export_complete_results_xlsx(request: Request, activity_id: int):
        """Export the complete roster and result in a WPS-friendly workbook."""

        require_session(request, "admin")
        connection = connect(config.database_path)
        try:
            connection.execute("BEGIN")
            activity = current_activity(connection)
            if int(activity["id"]) != activity_id:
                raise HTTPException(
                    status_code=409, detail="当前活动已经变化，请刷新页面后重试"
                )
            result_rows = connection.execute(
                """
                SELECT s.student_no, s.name, m.name AS major_name,
                       CASE WHEN se.id IS NULL THEN '未选' ELSE '已选' END
                         AS selection_status,
                       g.name AS group_name, se.selected_at
                FROM students s
                JOIN majors m ON m.id = s.major_id
                LEFT JOIN selections se
                  ON se.student_id = s.id AND se.revoked_at IS NULL
                LEFT JOIN teaching_groups g ON g.id = se.group_id
                WHERE s.active = 1
                ORDER BY m.sort_order, s.student_no
                """
            ).fetchall()
            group_rows = connection.execute(
                """
                SELECT g.name, g.total_capacity,
                       COUNT(se.id) AS selected_count
                FROM teaching_groups g
                LEFT JOIN selections se
                  ON se.group_id = g.id AND se.revoked_at IS NULL
                WHERE g.active = 1
                GROUP BY g.id, g.name, g.total_capacity, g.sort_order
                ORDER BY g.sort_order, g.id
                """
            ).fetchall()
            major_rows = connection.execute(
                """
                SELECT m.name,
                       COUNT(DISTINCT CASE WHEN s.active = 1 THEN s.id END)
                         AS student_count,
                       COUNT(DISTINCT CASE
                         WHEN s.active = 1 AND se.id IS NOT NULL THEN s.id END)
                         AS selected_count
                FROM majors m
                LEFT JOIN students s ON s.major_id = m.id
                LEFT JOIN selections se
                  ON se.student_id = s.id AND se.revoked_at IS NULL
                WHERE m.active = 1
                GROUP BY m.id, m.name, m.sort_order
                ORDER BY m.sort_order, m.id
                """
            ).fetchall()
        finally:
            connection.close()

        workbook = Workbook()
        workbook.iso_dates = True
        results_sheet = workbook.active
        results_sheet.title = "完整结果"
        summary_sheet = workbook.create_sheet("汇总")

        brand_fill = PatternFill("solid", fgColor="6E2432")
        accent_fill = PatternFill("solid", fgColor="E8D6C5")
        pale_fill = PatternFill("solid", fgColor="F7F1EC")
        header_font = Font(name="等线", size=11, bold=True, color="FFFFFF")
        body_font = Font(name="等线", size=11, color="2B2223")
        title_font = Font(name="等线", size=18, bold=True, color="6E2432")
        thin_side = Side(style="thin", color="D8C9C2")
        cell_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )
        centered = Alignment(horizontal="center", vertical="center")

        def xlsx_safe_text(value: Any) -> str:
            """Keep user-controlled workbook text from becoming a formula."""

            text = "" if value is None else str(value)
            visible = text.lstrip(" \t\r\n")
            if text.startswith(("\t", "\r", "\n")) or visible.startswith(
                ("=", "+", "-", "@")
            ):
                return "'" + text
            return text

        def xlsx_timestamp(value: Any) -> Any:
            if not value:
                return ""
            try:
                parsed = datetime.fromisoformat(str(value))
            except ValueError:
                return xlsx_safe_text(value)
            # Excel stores naive wall times.  Export Beijing local time because
            # this classroom application and its operators are in that zone.
            return parsed.astimezone(timezone(timedelta(hours=8))).replace(tzinfo=None)

        result_headers = ["学号", "姓名", "专业", "状态", "教学组", "选择时间"]
        results_sheet.append(result_headers)
        for row in result_rows:
            results_sheet.append(
                [
                    xlsx_safe_text(row["student_no"]),
                    xlsx_safe_text(row["name"]),
                    xlsx_safe_text(row["major_name"]),
                    row["selection_status"],
                    xlsx_safe_text(row["group_name"]),
                    xlsx_timestamp(row["selected_at"]),
                ]
            )
        for cell in results_sheet[1]:
            cell.fill = brand_fill
            cell.font = header_font
            cell.alignment = centered
            cell.border = cell_border
        for row in results_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.border = cell_border
                cell.alignment = Alignment(vertical="center")
            row[0].number_format = "@"
            row[0].alignment = Alignment(horizontal="left", vertical="center")
            row[3].alignment = centered
            if isinstance(row[5].value, datetime):
                row[5].number_format = "yyyy-mm-dd hh:mm:ss"
        results_sheet.freeze_panes = "A2"
        results_sheet.auto_filter.ref = f"A1:F{max(1, results_sheet.max_row)}"
        results_sheet.sheet_view.showGridLines = False
        results_sheet.row_dimensions[1].height = 26
        for column, width in {
            "A": 36,
            "B": 20,
            "C": 20,
            "D": 12,
            "E": 24,
            "F": 28,
        }.items():
            results_sheet.column_dimensions[column].width = width
        if result_rows:
            table = Table(displayName="CompleteResults", ref=f"A1:F{results_sheet.max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            results_sheet.add_table(table)

        total_students = len(result_rows)
        selected_students = sum(
            1 for row in result_rows if row["selection_status"] == "已选"
        )
        summary_sheet.merge_cells("A1:D1")
        summary_sheet["A1"] = "教学组抢选结果汇总"
        summary_sheet["A1"].font = title_font
        summary_sheet["A1"].alignment = Alignment(vertical="center")
        summary_sheet.row_dimensions[1].height = 34
        metadata = [
            ("活动名称", xlsx_safe_text(activity["title"])),
            ("活动编号", xlsx_safe_text(activity["code"])),
            ("导出时间", xlsx_timestamp(utc_now())),
        ]
        for row_index, (label, value) in enumerate(metadata, start=3):
            summary_sheet.cell(row_index, 1, label)
            summary_sheet.cell(row_index, 2, value)
            summary_sheet.cell(row_index, 1).fill = accent_fill
            summary_sheet.cell(row_index, 1).font = Font(
                name="等线", size=11, bold=True, color="6E2432"
            )

        summary_sheet.append([])
        metric_header_row = 7
        metrics = [
            ("总学生", total_students),
            ("已选", selected_students),
            ("未选", total_students - selected_students),
            (
                "完成率",
                selected_students / total_students if total_students else 0,
            ),
        ]
        for column, (label, value) in enumerate(metrics, start=1):
            label_cell = summary_sheet.cell(metric_header_row, column, label)
            value_cell = summary_sheet.cell(metric_header_row + 1, column, value)
            label_cell.fill = brand_fill
            label_cell.font = header_font
            label_cell.alignment = centered
            value_cell.fill = pale_fill
            value_cell.font = Font(name="等线", size=14, bold=True, color="6E2432")
            value_cell.alignment = centered
            if label == "完成率":
                value_cell.number_format = "0.0%"

        group_start = 11
        summary_sheet.cell(group_start, 1, "教学组")
        summary_sheet.cell(group_start, 2, "容量")
        summary_sheet.cell(group_start, 3, "已选")
        summary_sheet.cell(group_start, 4, "剩余")
        for cell in summary_sheet[group_start]:
            cell.fill = brand_fill
            cell.font = header_font
            cell.alignment = centered
        for row_index, row in enumerate(group_rows, start=group_start + 1):
            selected_count = int(row["selected_count"])
            summary_sheet.append(
                [
                    xlsx_safe_text(row["name"]),
                    int(row["total_capacity"]),
                    selected_count,
                    int(row["total_capacity"]) - selected_count,
                ]
            )

        major_start = group_start + len(group_rows) + 3
        for column, value in enumerate(["专业", "学生数", "已选", "未选"], start=1):
            cell = summary_sheet.cell(major_start, column, value)
            cell.fill = brand_fill
            cell.font = header_font
            cell.alignment = centered
        for row_index, row in enumerate(major_rows, start=major_start + 1):
            student_count = int(row["student_count"])
            selected_count = int(row["selected_count"])
            for column, value in enumerate(
                [
                    xlsx_safe_text(row["name"]),
                    student_count,
                    selected_count,
                    student_count - selected_count,
                ],
                start=1,
            ):
                summary_sheet.cell(row_index, column, value)

        for row in summary_sheet.iter_rows(
            min_row=3, max_row=summary_sheet.max_row, min_col=1, max_col=4
        ):
            for cell in row:
                if cell.value is not None:
                    cell.border = cell_border
                    if cell.font == Font():
                        cell.font = body_font
                    if cell.alignment == Alignment():
                        cell.alignment = Alignment(vertical="center")
        summary_sheet.freeze_panes = "A11"
        summary_sheet.sheet_view.showGridLines = False
        for column, width in {"A": 28, "B": 26, "C": 18, "D": 18}.items():
            summary_sheet.column_dimensions[column].width = width

        output = io.BytesIO()
        workbook.save(output)
        return Response(
            content=output.getvalue(),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": (
                    f'attachment; filename="{activity["code"]}-results.xlsx"'
                ),
                "Cache-Control": "no-store",
            },
        )

    @app.get("/api/admin/audit")
    def audit_log(request: Request, limit: int = 100):
        require_session(request, "admin")
        limit = min(max(limit, 1), 500)
        connection = connect(config.database_path)
        try:
            rows = connection.execute(
                """
                SELECT id, occurred_at, actor_type, actor_id, action,
                       entity_type, entity_id, details_json
                FROM audit_logs
                WHERE activity_id = (SELECT current_activity_id FROM settings WHERE id = 1)
                ORDER BY id DESC LIMIT ?
                """,
                (limit,),
            ).fetchall()
            return [
                {
                    **{key: row[key] for key in row.keys() if key != "details_json"},
                    "details": json.loads(row["details_json"]),
                }
                for row in rows
            ]
        finally:
            connection.close()

    @app.get("/api/admin/qr.png")
    def qr_code(request: Request):
        require_session(request, "admin")
        connection = connect(config.database_path)
        try:
            settings = setting_dict(connection)
        finally:
            connection.close()
        base_url = settings["public_base_url"] or config.public_base_url
        if not base_url:
            raise HTTPException(status_code=409, detail="请先在系统设置中填写学生端访问地址")
        qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=12, border=4)
        qr.add_data(base_url.rstrip("/") + "/")
        qr.make(fit=True)
        image = qr.make_image(fill_color="#2b2223", back_color="#ffffff")
        output = io.BytesIO()
        image.save(output, format="PNG")
        output.seek(0)
        return StreamingResponse(output, media_type="image/png", headers={"Cache-Control": "no-store"})

    @app.get("/brand/college-wordmark-official.png", include_in_schema=False)
    def brand_asset():
        return FileResponse(
            BRAND_ROOT / "college-wordmark-official.png",
            media_type="image/png",
            headers={"Cache-Control": "public, max-age=86400"},
        )

    @app.get("/assets/app.css", include_in_schema=False)
    def css_asset():
        return FileResponse(WEB_ROOT / "app.css", media_type="text/css")

    @app.get("/assets/student.js", include_in_schema=False)
    def student_script():
        return FileResponse(WEB_ROOT / "student.js", media_type="text/javascript")

    @app.get("/assets/admin.js", include_in_schema=False)
    def admin_script():
        return FileResponse(WEB_ROOT / "admin.js", media_type="text/javascript")

    @app.get("/admin", include_in_schema=False)
    def admin_page():
        return FileResponse(WEB_ROOT / "admin.html", media_type="text/html")

    @app.get("/", include_in_schema=False)
    def student_page():
        return FileResponse(WEB_ROOT / "index.html", media_type="text/html")

    return app
