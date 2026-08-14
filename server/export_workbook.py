from __future__ import annotations

import io
import math
import unicodedata
from datetime import datetime, timedelta, timezone
from typing import Any, Iterable, Literal, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ExportKind = Literal["complete", "selections", "unselected"]

BRAND_COLOR = "6E2432"
ACCENT_COLOR = "E8D6C5"
PALE_COLOR = "F7F1EC"
TEXT_COLOR = "2B2223"
BORDER_COLOR = "D8C9C2"
BEIJING_TIMEZONE = timezone(timedelta(hours=8))
SELECTION_SOURCE_LABELS = {
    "student": "学生提交",
    "admin": "管理员补位",
}


def spreadsheet_safe_text(value: Any) -> str:
    """Return text that spreadsheet applications cannot interpret as a formula."""

    text = "" if value is None else str(value)
    visible = text.lstrip(" \t\r\n")
    if text.startswith(("\t", "\r", "\n")) or visible.startswith(
        ("=", "+", "-", "@")
    ):
        return "'" + text
    return text


def spreadsheet_timestamp(value: Any) -> Any:
    """Convert an ISO timestamp to a timezone-free Beijing wall time for Excel."""

    if not value:
        return ""
    try:
        parsed = datetime.fromisoformat(str(value))
    except ValueError:
        return spreadsheet_safe_text(value)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(BEIJING_TIMEZONE).replace(tzinfo=None)


def _display_units(value: Any) -> int:
    """Approximate the width WPS/Excel needs for mixed Chinese and Latin text."""

    return sum(
        2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
        for character in str(value or "")
    )


def _fit_column_widths(
    sheet: Any,
    *,
    minimums: Sequence[float],
    maximums: Sequence[float],
) -> tuple[float, ...]:
    widths: list[float] = []
    for column_index, (minimum, maximum) in enumerate(
        zip(minimums, maximums, strict=True), start=1
    ):
        required = max(
            (_display_units(sheet.cell(row_index, column_index).value) + 2)
            for row_index in range(1, sheet.max_row + 1)
        )
        widths.append(max(minimum, min(maximum, float(required))))
    return tuple(widths)


def _wrapped_line_count(value: Any, column_width: float) -> int:
    available_units = max(1, int(column_width) - 2)
    return max(
        1,
        sum(
            max(1, math.ceil(_display_units(line) / available_units))
            for line in str(value or "").splitlines() or [""]
        ),
    )


def _style_table_sheet(
    sheet: Any,
    *,
    column_widths: Sequence[float],
    table_name: str,
    timestamp_columns: Sequence[int] = (),
    wrap_columns: Sequence[int] = (),
) -> None:
    brand_fill = PatternFill("solid", fgColor=BRAND_COLOR)
    header_font = Font(name="等线", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="等线", size=10, color=TEXT_COLOR)
    thin_side = Side(style="thin", color=BORDER_COLOR)
    cell_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    centered = Alignment(horizontal="center", vertical="center")
    wrapped_centered = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    wrap_column_set = set(wrap_columns)

    for cell in sheet[1]:
        cell.fill = brand_fill
        cell.font = header_font
        cell.alignment = centered
        cell.border = cell_border
    for row in sheet.iter_rows(min_row=2):
        row_index = row[0].row
        wrapped_lines = 1
        for column_index, cell in enumerate(row, start=1):
            cell.font = body_font
            cell.border = cell_border
            if column_index in wrap_column_set:
                cell.alignment = wrapped_centered
                wrapped_lines = max(
                    wrapped_lines,
                    _wrapped_line_count(cell.value, column_widths[column_index - 1]),
                )
            else:
                cell.alignment = centered
        row[0].number_format = "@"
        for column_index in timestamp_columns:
            timestamp_cell = row[column_index - 1]
            if isinstance(timestamp_cell.value, datetime):
                timestamp_cell.number_format = "yyyy-mm-dd hh:mm:ss"
        sheet.row_dimensions[row_index].height = max(22, 15 * wrapped_lines + 7)

    last_column = sheet.cell(row=1, column=len(column_widths)).column_letter
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{last_column}{max(1, sheet.max_row)}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 26
    for column_index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[
            sheet.cell(row=1, column=column_index).column_letter
        ].width = width

    if sheet.max_row > 1:
        table = Table(
            displayName=table_name,
            ref=f"A1:{last_column}{sheet.max_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_options.horizontalCentered = True
    sheet.print_title_rows = "1:1"
    sheet.print_area = f"A1:{last_column}{max(1, sheet.max_row)}"
    sheet.page_margins.left = 0.3
    sheet.page_margins.right = 0.3
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.2
    sheet.page_margins.footer = 0.2
    sheet.oddHeader.center.text = (
        f"&B安徽建筑大学 · 建筑与空间规划学院  {sheet.title}"
    )
    sheet.oddFooter.center.text = "第 &P 页 / 共 &N 页"
    sheet.oddFooter.right.text = "制作：Mikutea"


def _append_summary_sheet(
    workbook: Workbook,
    *,
    activity: Mapping[str, Any],
    exported_at: str,
    result_rows: Sequence[Mapping[str, Any]],
    group_rows: Sequence[Mapping[str, Any]],
    major_rows: Sequence[Mapping[str, Any]],
    kind: ExportKind,
) -> None:
    sheet = workbook.create_sheet("汇总")
    brand_fill = PatternFill("solid", fgColor=BRAND_COLOR)
    accent_fill = PatternFill("solid", fgColor=ACCENT_COLOR)
    pale_fill = PatternFill("solid", fgColor=PALE_COLOR)
    header_font = Font(name="等线", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="等线", size=10, color=TEXT_COLOR)
    title_font = Font(name="等线", size=18, bold=True, color=BRAND_COLOR)
    thin_side = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    centered = Alignment(horizontal="center", vertical="center")
    wrapped_centered = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    sheet.merge_cells("A1:D1")
    sheet["A1"] = {
        "complete": "教学组抢选结果汇总",
        "selections": "教学组抢选选择记录汇总",
        "unselected": "教学组抢选未选名单汇总",
    }[kind]
    sheet["A1"].font = title_font
    sheet["A1"].alignment = centered
    sheet.row_dimensions[1].height = 34

    metadata = [
        ("活动名称", spreadsheet_safe_text(activity["title"])),
        ("活动编号", spreadsheet_safe_text(activity["code"])),
        ("导出时间", spreadsheet_timestamp(exported_at)),
    ]
    for row_index, (label, value) in enumerate(metadata, start=3):
        label_cell = sheet.cell(row_index, 1, label)
        value_cell = sheet.cell(row_index, 2, value)
        label_cell.fill = accent_fill
        label_cell.font = Font(name="等线", size=10, bold=True, color=BRAND_COLOR)
        for cell in (label_cell, value_cell):
            cell.border = border
            cell.alignment = centered
        value_cell.font = body_font
        value_cell.alignment = wrapped_centered
        sheet.row_dimensions[row_index].height = max(
            22, 15 * _wrapped_line_count(value, 28) + 7
        )
        if isinstance(value, datetime):
            value_cell.number_format = "yyyy-mm-dd hh:mm:ss"

    total_students = len(result_rows)
    selected_students = sum(
        1 for row in result_rows if row["selection_status"] == "已选"
    )
    metrics = [
        ("总学生", total_students),
        ("已选", selected_students),
        ("未选", total_students - selected_students),
        ("完成率", selected_students / total_students if total_students else 0),
    ]
    for column, (label, value) in enumerate(metrics, start=1):
        label_cell = sheet.cell(7, column, label)
        value_cell = sheet.cell(8, column, value)
        label_cell.fill = brand_fill
        label_cell.font = header_font
        value_cell.fill = pale_fill
        value_cell.font = Font(name="等线", size=14, bold=True, color=BRAND_COLOR)
        for cell in (label_cell, value_cell):
            cell.border = border
            cell.alignment = centered
        if label == "完成率":
            value_cell.number_format = "0.0%"

    group_start = 11
    for column, value in enumerate(["教学组", "容量", "已选", "剩余"], start=1):
        cell = sheet.cell(group_start, column, value)
        cell.fill = brand_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = centered
    for row_index, row in enumerate(group_rows, start=group_start + 1):
        selected_count = int(row["selected_count"])
        values = [
            spreadsheet_safe_text(row["name"]),
            int(row["total_capacity"]),
            selected_count,
            int(row["total_capacity"]) - selected_count,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = body_font
            cell.border = border
            cell.alignment = wrapped_centered if column == 1 else centered
        sheet.row_dimensions[row_index].height = max(
            22, 15 * _wrapped_line_count(values[0], 30) + 7
        )

    major_start = group_start + len(group_rows) + 3
    for column, value in enumerate(["专业", "学生数", "已选", "未选"], start=1):
        cell = sheet.cell(major_start, column, value)
        cell.fill = brand_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = centered
    for row_index, row in enumerate(major_rows, start=major_start + 1):
        student_count = int(row["student_count"])
        selected_count = int(row["selected_count"])
        values = [
            spreadsheet_safe_text(row["name"]),
            student_count,
            selected_count,
            student_count - selected_count,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = body_font
            cell.border = border
            cell.alignment = wrapped_centered if column == 1 else centered
        sheet.row_dimensions[row_index].height = max(
            22, 15 * _wrapped_line_count(values[0], 30) + 7
        )

    sheet.freeze_panes = "A11"
    sheet.sheet_view.showGridLines = False
    for column, width in {"A": 30, "B": 28, "C": 14, "D": 14}.items():
        sheet.column_dimensions[column].width = width
    for row_index in range(3, sheet.max_row + 1):
        if sheet.row_dimensions[row_index].height is None:
            sheet.row_dimensions[row_index].height = 22

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_options.horizontalCentered = True
    sheet.print_title_rows = "1:1"
    sheet.print_area = f"A1:D{sheet.max_row}"
    sheet.page_margins.left = 0.4
    sheet.page_margins.right = 0.4
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55
    sheet.page_margins.header = 0.2
    sheet.page_margins.footer = 0.2
    sheet.oddHeader.center.text = "&B安徽建筑大学 · 建筑与空间规划学院"
    sheet.oddFooter.center.text = "第 &P 页 / 共 &N 页"
    sheet.oddFooter.right.text = "制作：Mikutea"


def build_export_workbook(
    *,
    activity: Mapping[str, Any],
    exported_at: str,
    kind: ExportKind,
    data_rows: Iterable[Mapping[str, Any]],
    result_rows: Iterable[Mapping[str, Any]],
    group_rows: Iterable[Mapping[str, Any]],
    major_rows: Iterable[Mapping[str, Any]],
) -> bytes:
    """Build a print-ready complete, selected, or unselected WPS workbook."""

    data = list(data_rows)
    results = list(result_rows)
    groups = list(group_rows)
    majors = list(major_rows)
    workbook = Workbook()
    workbook.iso_dates = True
    sheet = workbook.active

    if kind == "complete":
        sheet.title = "完整结果"
        sheet.append(["学号", "姓名", "专业", "状态", "教学组", "选择时间"])
        for row in data:
            sheet.append(
                [
                    spreadsheet_safe_text(row["student_no"]),
                    spreadsheet_safe_text(row["name"]),
                    spreadsheet_safe_text(row["major_name"]),
                    row["selection_status"],
                    spreadsheet_safe_text(row["group_name"]),
                    spreadsheet_timestamp(row["selected_at"]),
                ]
            )
        widths = _fit_column_widths(
            sheet,
            minimums=(15, 12, 14, 10, 14, 19),
            maximums=(15, 20, 26, 10, 26, 21),
        )
        wrap_columns = (2, 3, 5)
        table_name = "CompleteResults"
        timestamp_columns = (6,)
    elif kind == "selections":
        sheet.title = "选择记录"
        sheet.append(["学号", "姓名", "专业", "教学组", "选择时间", "来源"])
        for row in data:
            sheet.append(
                [
                    spreadsheet_safe_text(row["student_no"]),
                    spreadsheet_safe_text(row["name"]),
                    spreadsheet_safe_text(row["major_name"]),
                    spreadsheet_safe_text(row["group_name"]),
                    spreadsheet_timestamp(row["selected_at"]),
                    spreadsheet_safe_text(
                        SELECTION_SOURCE_LABELS.get(str(row["source"]), row["source"])
                    ),
                ]
            )
        widths = _fit_column_widths(
            sheet,
            minimums=(15, 12, 14, 14, 19, 12),
            maximums=(15, 20, 26, 26, 21, 12),
        )
        wrap_columns = (2, 3, 4)
        table_name = "SuccessfulSelections"
        timestamp_columns = (5,)
    else:
        sheet.title = "未选名单"
        sheet.append(["学号", "姓名", "专业"])
        for row in data:
            sheet.append(
                [
                    spreadsheet_safe_text(row["student_no"]),
                    spreadsheet_safe_text(row["name"]),
                    spreadsheet_safe_text(row["major_name"]),
                ]
            )
        widths = _fit_column_widths(
            sheet,
            minimums=(15, 12, 14),
            maximums=(15, 20, 28),
        )
        wrap_columns = (2, 3)
        table_name = "UnselectedStudents"
        timestamp_columns = ()

    _style_table_sheet(
        sheet,
        column_widths=widths,
        table_name=table_name,
        timestamp_columns=timestamp_columns,
        wrap_columns=wrap_columns,
    )
    _append_summary_sheet(
        workbook,
        activity=activity,
        exported_at=exported_at,
        result_rows=results,
        group_rows=groups,
        major_rows=majors,
        kind=kind,
    )
    workbook.active = 0
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
